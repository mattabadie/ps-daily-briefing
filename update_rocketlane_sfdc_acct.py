#!/usr/bin/env python3
"""
update_rocketlane_sfdc_acct.py

Backfill the "SFDC Acct No" custom field on Rocketlane company records,
using the matched Salesforce Account IDs from
  Exterro/outputs/Rocketlane_to_SF_Account_Match.xlsx

Default: DRY RUN — prints exactly what would happen, makes ZERO API writes.
Pass --execute to actually push updates.

Recommended order of operations:
  1. python3 update_rocketlane_sfdc_acct.py
       → dry-run, shows the proposed updates
  2. python3 update_rocketlane_sfdc_acct.py --test-one 526735
       → updates ONE company (BankUnited) so you can verify in Rocketlane UI
  3. python3 update_rocketlane_sfdc_acct.py --execute
       → bulk update all confident matches

Match-type filter (Match Type column in the spreadsheet):
  --include-exact      include rows marked "exact"     (DEFAULT: yes)
  --include-fuzzy      include rows marked "fuzzy"     (DEFAULT: yes)
  --include-ambiguous  include rows marked "ambiguous" (DEFAULT: no — needs human review)

Every successful update is appended to:
  Exterro/outputs/sfdc_acct_backfill_log.csv
"""
import argparse
import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from rocketlane_client import (
    api_get,
    api_request,
    V1_BASE_URL,
    extract_acct_id_from_fields,
    SF_ACCT_ID_RE,
    API_KEY,
)

# ─── Paths ─────────────────────────────────────────────────────────────────────
WORKSPACE = Path(
    "/Users/matthew.abadie/Library/Mobile Documents/com~apple~CloudDocs/"
    "iCloud Storage/Exterro"
)
DEFAULT_INPUT = WORKSPACE / "outputs" / "Rocketlane_to_SF_Account_Match.xlsx"
LOG_FILE = WORKSPACE / "outputs" / "sfdc_acct_backfill_log.csv"

# Companies known to have SFDC Acct No populated — used to discover the fieldId.
# Verified from probe_company_full.py output. ServiceNow has 3 fields, of which
# fields[2] holds the SF Acct ID (matches 001-prefix regex).
KNOWN_POPULATED_COMPANY_IDS = [478735, 483041, 483056]  # ServiceNow, McKinsey, Xcel


# ═══════════════════════════════════════════════════════════════════════════════
# Discover SFDC Acct No fieldId from a known-populated company
# ═══════════════════════════════════════════════════════════════════════════════
def discover_sfdc_acct_field_id():
    """Look at a known-populated company; find the field whose value matches a
    Salesforce Account ID pattern. Returns the fieldId or raises."""
    print("Discovering SFDC Acct No fieldId from a known-populated company...")
    for cid in KNOWN_POPULATED_COMPANY_IDS:
        try:
            co = api_get(f"companies/{cid}", base=V1_BASE_URL, timeout=60)
        except Exception as e:
            print(f"  WARN companyId={cid}: {str(e)[:80]}")
            continue
        for f in co.get("fields") or []:
            val = str(f.get("fieldValue") or "")
            if SF_ACCT_ID_RE.search(val):
                fid = f.get("fieldId")
                co_name = co.get("companyName")
                print(f"  ✓ Found fieldId={fid} via '{co_name}' (companyId={cid})")
                print(f"    sample value: {val!r}")
                return fid
    raise RuntimeError(
        "Could not discover SFDC Acct No fieldId — none of the seed companies "
        "had a 001-prefixed value. Add a known-populated companyId to "
        "KNOWN_POPULATED_COMPANY_IDS."
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Read matches from the spreadsheet
# ═══════════════════════════════════════════════════════════════════════════════
def read_matches(xlsx_path):
    """Returns list of dicts with keys: rl_id, rl_name, sf_id, sf_name, match_type."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Rocketlane to SF Account"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rl_id = r[idx["Rocketlane Company ID"]]
        rl_name = r[idx["Rocketlane Name"]]
        sf_id = r[idx["SF Account ID (SFDC Acct No)"]]
        sf_name = r[idx["SF Account Name"]]
        match_type = r[idx["Match Type"]]
        rows.append({
            "rl_id": rl_id,
            "rl_name": rl_name,
            "sf_id": (sf_id or "").strip() if sf_id else "",
            "sf_name": sf_name or "",
            "match_type": match_type or "",
        })
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# Update one company's SFDC Acct No field
# ═══════════════════════════════════════════════════════════════════════════════
def update_company_field(company_id, field_id, new_value, *, dry_run=True):
    """PATCH the company's custom field. Try a few likely body shapes — the
    Rocketlane v1 update endpoint isn't documented publicly so we attempt the
    most common pattern first.

    Returns (success: bool, response_or_error: dict|str)."""
    if dry_run:
        return True, {"dryrun": True, "would_set": new_value}

    # Attempt 1 — PATCH /v1/companies/{id} with {"fields": [{"fieldId":X,"fieldValue":Y}]}
    body = {"fields": [{"fieldId": field_id, "fieldValue": new_value}]}
    try:
        resp = api_request("PATCH", f"companies/{company_id}", body=body,
                           base=V1_BASE_URL, timeout=60)
        return True, resp
    except RuntimeError as e1:
        err1 = str(e1)
        # Attempt 2 — PUT same body
        try:
            resp = api_request("PUT", f"companies/{company_id}", body=body,
                               base=V1_BASE_URL, timeout=60)
            return True, resp
        except RuntimeError as e2:
            err2 = str(e2)
            return False, f"PATCH failed: {err1[:200]}; PUT failed: {err2[:200]}"


def verify_update(company_id, expected_value):
    """Re-read the company; confirm its fields contain the expected acct id."""
    try:
        co = api_get(f"companies/{company_id}", base=V1_BASE_URL, timeout=60)
    except Exception as e:
        return False, f"verify fetch failed: {e}"
    found = extract_acct_id_from_fields(co.get("fields") or [])
    if found and found.lower() == expected_value.lower():
        return True, found
    return False, f"expected {expected_value}, got {found!r}"


# ═══════════════════════════════════════════════════════════════════════════════
# Logging
# ═══════════════════════════════════════════════════════════════════════════════
def log_change(rl_id, rl_name, sf_id, sf_name, match_type, status, detail):
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    new_file = not LOG_FILE.exists()
    with LOG_FILE.open("a", newline="") as fp:
        w = csv.writer(fp)
        if new_file:
            w.writerow(["timestamp", "rl_id", "rl_name", "sf_id", "sf_name",
                        "match_type", "status", "detail"])
        w.writerow([
            datetime.now().isoformat(timespec="seconds"),
            rl_id, rl_name, sf_id, sf_name, match_type, status, str(detail)[:200],
        ])


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=str(DEFAULT_INPUT),
                    help="Match spreadsheet path (xlsx)")
    ap.add_argument("--execute", action="store_true",
                    help="Actually push updates. Default is dry-run.")
    ap.add_argument("--test-one", type=int, metavar="COMPANY_ID",
                    help="Update ONE company by Rocketlane ID, then exit. Forces --execute.")
    ap.add_argument("--include-exact", dest="inc_exact", action="store_true", default=True)
    ap.add_argument("--no-exact", dest="inc_exact", action="store_false")
    ap.add_argument("--include-fuzzy", dest="inc_fuzzy", action="store_true", default=True)
    ap.add_argument("--no-fuzzy", dest="inc_fuzzy", action="store_false")
    ap.add_argument("--include-ambiguous", dest="inc_amb", action="store_true", default=False,
                    help="Also include 'ambiguous' rows (best-guess; risky — review xlsx first)")
    args = ap.parse_args()

    if not API_KEY:
        print("ERROR: ROCKETLANE_API_KEY not set in env or rocketlane.env", file=sys.stderr)
        sys.exit(1)

    if not Path(args.input).exists():
        print(f"ERROR: input not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    if args.test_one:
        args.execute = True

    print("=" * 72)
    print(f"Rocketlane SFDC Acct No backfill {'[EXECUTE]' if args.execute else '[DRY RUN]'}")
    print(f"Input: {args.input}")
    print("=" * 72)

    field_id = discover_sfdc_acct_field_id()
    print()

    rows = read_matches(args.input)

    # Filter
    selected_types = set()
    if args.inc_exact:    selected_types.add("exact")
    if args.inc_fuzzy:    selected_types.add("fuzzy")
    if args.inc_amb:      selected_types.add("ambiguous")

    candidates = [r for r in rows if r["match_type"] in selected_types and r["sf_id"]]

    if args.test_one:
        candidates = [r for r in candidates if r["rl_id"] == args.test_one]
        if not candidates:
            print(f"ERROR: companyId {args.test_one} not in match set "
                  f"(or its match_type is excluded).")
            sys.exit(1)

    print(f"Candidates to update: {len(candidates)}")
    if not candidates:
        print("Nothing to do.")
        return

    # Show summary
    by_type = {}
    for c in candidates:
        by_type.setdefault(c["match_type"], 0)
        by_type[c["match_type"]] += 1
    for t, n in sorted(by_type.items()):
        print(f"  {t}: {n}")

    if not args.execute:
        print()
        print("DRY RUN — first 10 proposed updates:")
        for c in candidates[:10]:
            print(f"  {c['rl_id']:>8}  {c['rl_name'][:35]:35s}  →  {c['sf_id']}  ({c['match_type']})")
        if len(candidates) > 10:
            print(f"  ... and {len(candidates) - 10} more")
        print()
        print("To actually push: re-run with --execute")
        print("To test on ONE record first: --test-one <companyId>")
        return

    # Real push
    print()
    print(f"Pushing {len(candidates)} updates (writing to log: {LOG_FILE})...")
    ok, fail = 0, 0
    for i, c in enumerate(candidates, 1):
        success, detail = update_company_field(c["rl_id"], field_id, c["sf_id"],
                                               dry_run=False)
        if success:
            v_ok, v_detail = verify_update(c["rl_id"], c["sf_id"])
            if v_ok:
                print(f"  [{i}/{len(candidates)}] ✓ {c['rl_name'][:40]:40s} → {c['sf_id']}")
                log_change(c["rl_id"], c["rl_name"], c["sf_id"], c["sf_name"],
                           c["match_type"], "OK", v_detail)
                ok += 1
            else:
                print(f"  [{i}/{len(candidates)}] ⚠ {c['rl_name'][:40]:40s} update returned OK "
                      f"but verify failed: {v_detail}")
                log_change(c["rl_id"], c["rl_name"], c["sf_id"], c["sf_name"],
                           c["match_type"], "VERIFY_FAILED", v_detail)
                fail += 1
        else:
            print(f"  [{i}/{len(candidates)}] ✗ {c['rl_name'][:40]:40s} {detail[:120]}")
            log_change(c["rl_id"], c["rl_name"], c["sf_id"], c["sf_name"],
                       c["match_type"], "FAILED", detail)
            fail += 1

        if args.test_one:
            break  # only one update in test mode

    print()
    print("=" * 72)
    print(f"DONE — {ok} succeeded, {fail} failed")
    print(f"Log: {LOG_FILE}")
    print("=" * 72)


if __name__ == "__main__":
    main()
