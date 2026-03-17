#!/usr/bin/env python3
"""
Subscription Projects Audit
Generates an Excel audit of all active subscription projects under Post Implementation,
including contract type, budget data, and flags for projects needing correction.

Outputs:
  - Excel file emailed to VP
  - Summary posted to Google Chat

Usage:
  python subscription_audit.py                # Generate and email audit
  python subscription_audit.py --dry-run      # Preview without sending
"""

import argparse
import json
import os
import re
import smtplib
import sys
import time
import urllib.error
import urllib.request
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
API_KEY = os.environ.get("ROCKETLANE_API_KEY", "")
GMAIL_ADDRESS = os.environ.get("GMAIL_ADDRESS", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
GCHAT_WEBHOOK_URL = os.environ.get("GCHAT_SUB_WEBHOOK_URL", "")
BASE_URL = "https://services.api.exterro.com/api/1.0"
RL_APP_BASE = "https://services.exterro.com/projects"

ORONDE_ID = 393607
ACTIVE_STATUS_VALUES = {2, 4, 5, 6, 9, 12, 14, 15}
NOW = datetime.now()
DASH = "\u2014"

EXTRA_RECIPIENTS = ["matt.abadie@exterro.com"]


# ═══════════════════════════════════════════════════════════════════════════════
# HTTP + DATA FETCHING
# ═══════════════════════════════════════════════════════════════════════════════
def api_get(path, retries=3):
    url = f"{BASE_URL}/{path}"
    req = urllib.request.Request(url, headers={"api-key": API_KEY, "accept": "application/json"})
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < retries - 1:
                time.sleep(2 ** (attempt + 1))
            else:
                raise


def fetch_all_projects():
    all_projects, page_token = [], None
    while True:
        url = "projects" + (f"?pageToken={page_token}" if page_token else "")
        resp = api_get(url)
        all_projects.extend(resp.get("data", []))
        pag = resp.get("pagination", {})
        if pag.get("hasMore") and pag.get("nextPageToken"):
            page_token = pag["nextPageToken"]
        else:
            break
    return all_projects


def fetch_project_detail(pid):
    return api_get(f"projects/{pid}")


# ═══════════════════════════════════════════════════════════════════════════════
# FIELD HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def get_field(project, label):
    for f in project.get("fields", []):
        if f.get("fieldLabel") == label:
            return f.get("fieldValueLabel", f.get("fieldValue", ""))
    return None


def is_post_impl(p):
    member_ids = {m.get("userId") for m in p.get("teamMembers", {}).get("members", [])}
    return ORONDE_ID in member_ids or p.get("owner", {}).get("userId") == ORONDE_ID


def is_active_subscription(p):
    sv = p.get("status", {}).get("value")
    if sv not in ACTIVE_STATUS_VALUES:
        return False
    return (get_field(p, "Project Type") or "").lower() == "subscription"


# ═══════════════════════════════════════════════════════════════════════════════
# DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════
def extract_project_row(p, detail):
    pid = p.get("projectId", "")
    financials = detail.get("financials", {}) or {}
    contract_type = financials.get("contractType", "UNKNOWN") or "UNKNOWN"
    sub_contract = financials.get("subscriptionContract", {}) or {}
    tm_contract = financials.get("timeAndMaterialContract", {}) or {}

    period_minutes = sub_contract.get("periodMinutes", 0) or 0
    no_of_periods = sub_contract.get("noOfPeriods", 0) or 0
    total_minutes = period_minutes * no_of_periods
    total_hours = total_minutes / 60 if total_minutes else 0
    period_budget = sub_contract.get("periodBudget", 0) or 0
    frequency = sub_contract.get("subscriptionFrequency", "") or ""
    start_date = sub_contract.get("subscriptionStartDate", "") or ""
    tm_budget = tm_contract.get("projectBudget", 0) or 0

    owner = p.get("owner", {})
    owner_name = f'{owner.get("firstName", "")} {owner.get("lastName", "")}'.strip()
    customer = p.get("customer", {}).get("companyName", "N/A")

    hrs_match = re.search(r'(\d+)\s*(?:hrs?|hours?)', p.get("projectName", ""), re.IGNORECASE)

    return {
        "customer": customer,
        "project_name": p.get("projectName", "?"),
        "project_id": pid,
        "status": p.get("status", {}).get("label", "Unknown"),
        "pm": owner_name,
        "contract_type": contract_type,
        "frequency": frequency,
        "periods": no_of_periods,
        "period_hrs": round(period_minutes / 60, 1) if period_minutes else 0,
        "total_budget_hrs": round(total_hours, 1),
        "period_budget_usd": period_budget,
        "tm_budget_usd": tm_budget,
        "hrs_in_name": int(hrs_match.group(1)) if hrs_match else None,
        "start_date": start_date[:10] if start_date else "",
        "health": (get_field(p, "Red/Yellow/Green Health") or "").strip(),
        "domain": get_field(p, "Opp: Service Hours Domain(s)") or "",
        "opp_owner": get_field(p, "Opp: Opportunity Owner") or "",
        "client_segment": get_field(p, "Opp: Client Segmentation") or "",
        "needs_fix": contract_type not in ("SUBSCRIPTION", "UNKNOWN"),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscription Projects"

    headers = [
        "Customer", "Project Name", "Project ID", "Status", "PM",
        "Contract Type", "Frequency", "Periods", "Period Hrs",
        "Total Budget Hrs", "Period Budget ($)", "T&M Budget ($)",
        "Hrs in Name", "Start Date", "Health", "Domain",
        "Opp Owner", "Client Segment", "Needs Fix"
    ]

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    data_font = Font(name="Arial", size=10)
    money_fmt = '$#,##0'

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, 2):
        vals = [
            r["customer"], r["project_name"], r["project_id"], r["status"], r["pm"],
            r["contract_type"], r["frequency"], r["periods"], r["period_hrs"],
            r["total_budget_hrs"], r["period_budget_usd"], r["tm_budget_usd"],
            r["hrs_in_name"], r["start_date"], r["health"], r["domain"],
            r["opp_owner"], r["client_segment"], "YES" if r["needs_fix"] else ""
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = data_font
            c.border = thin_border
            if col in (11, 12):
                c.number_format = money_fmt
            if r["needs_fix"]:
                c.fill = red_fill
            if col == 15:
                h_lower = (v or "").lower()
                if "red" in h_lower:
                    c.fill = PatternFill("solid", fgColor="FCA5A5")
                elif "yellow" in h_lower:
                    c.fill = yellow_fill
                elif "green" in h_lower:
                    c.fill = green_fill

    widths = [28, 50, 12, 14, 20, 18, 12, 8, 10, 14, 14, 14, 10, 12, 10, 24, 22, 18, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = "1E293B"
    ct_counts = Counter(r["contract_type"] for r in rows)

    ws2["A1"] = "Contract Type Breakdown"
    ws2["A1"].font = Font(name="Arial", bold=True, size=12)
    ws2["A3"] = "Contract Type"
    ws2["B3"] = "Count"
    ws2["A3"].font = Font(name="Arial", bold=True, size=10)
    ws2["B3"].font = Font(name="Arial", bold=True, size=10)
    for i, (ct, cnt) in enumerate(sorted(ct_counts.items()), 4):
        ws2.cell(row=i, column=1, value=ct).font = data_font
        ws2.cell(row=i, column=2, value=cnt).font = data_font
        if ct not in ("SUBSCRIPTION", "UNKNOWN"):
            ws2.cell(row=i, column=1).fill = red_fill
            ws2.cell(row=i, column=2).fill = red_fill

    off = len(ct_counts) + 6
    ws2.cell(row=off, column=1, value="Budget Statistics").font = Font(name="Arial", bold=True, size=12)
    stats = [
        ("Total projects", len(rows)),
        ("With subscription budget", sum(1 for r in rows if r["total_budget_hrs"] > 0)),
        ("Need contract type fix", sum(1 for r in rows if r["needs_fix"])),
        ("Missing budget (correct type)", sum(1 for r in rows if r["total_budget_hrs"] == 0 and not r["needs_fix"])),
    ]
    for i, (label, val) in enumerate(stats, off + 2):
        ws2.cell(row=i, column=1, value=label).font = data_font
        c = ws2.cell(row=i, column=2, value=val)
        c.font = data_font
        if "fix" in label.lower():
            c.fill = red_fill

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    return wb


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL + CHAT
# ═══════════════════════════════════════════════════════════════════════════════
def send_audit_email(wb, rows, dry_run=False):
    total = len(rows)
    needs_fix = sum(1 for r in rows if r["needs_fix"])
    with_budget = sum(1 for r in rows if r["total_budget_hrs"] > 0)
    date_str = NOW.strftime("%b %d, %Y")

    subject = f"Subscription Projects Audit {DASH} {date_str} ({total} projects, {needs_fix} need correction)"

    body = f"""<html><body style="font-family: Arial, sans-serif; color: #1e293b;">
<h2 style="color: #1e293b;">Subscription Projects Audit</h2>
<p style="color: #64748b;">{date_str}</p>
<table style="border-collapse: collapse; font-size: 14px;">
<tr><td style="padding: 4px 12px; color: #64748b;">Total projects</td><td style="padding: 4px 12px; font-weight: bold;">{total}</td></tr>
<tr><td style="padding: 4px 12px; color: #64748b;">With subscription budget</td><td style="padding: 4px 12px; font-weight: bold;">{with_budget}</td></tr>
<tr><td style="padding: 4px 12px; color: #64748b;">Need contract type fix</td><td style="padding: 4px 12px; font-weight: bold; color: #ef4444;">{needs_fix}</td></tr>
</table>
<p style="font-size: 13px; color: #64748b; margin-top: 16px;">Full audit spreadsheet attached.</p>
<p style="font-size: 11px; color: #94a3b8; margin-top: 24px;">Auto-generated Subscription Audit &middot; Rocketlane API</p>
</body></html>"""

    recipients = set()
    recipients.add(GMAIL_ADDRESS)
    for r in EXTRA_RECIPIENTS:
        if r.strip():
            recipients.add(r.strip())
    to_list = ", ".join(sorted(recipients))

    if dry_run:
        print(f"  [DRY RUN] Would email to: {to_list}")
        print(f"  [DRY RUN] Subject: {subject}")
        return

    msg = MIMEMultipart("mixed")
    msg["From"] = GMAIL_ADDRESS
    msg["To"] = to_list
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    # Attach xlsx
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Subscription_Audit_{NOW.strftime('%Y%m%d')}.xlsx"
    attachment = MIMEApplication(buf.read(), Name=filename)
    attachment["Content-Disposition"] = f'attachment; filename="{filename}"'
    msg.attach(attachment)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.send_message(msg)
    print(f"  Email sent to: {to_list}")


def post_chat_summary(rows, dry_run=False):
    if not GCHAT_WEBHOOK_URL:
        print("  GCHAT_SUB_WEBHOOK_URL not set, skipping chat post.")
        return

    total = len(rows)
    needs_fix = [r for r in rows if r["needs_fix"]]
    with_budget = sum(1 for r in rows if r["total_budget_hrs"] > 0)
    date_str = NOW.strftime("%A, %B %d, %Y")

    ct_counts = Counter(r["contract_type"] for r in rows)
    ct_lines = []
    for ct, cnt in sorted(ct_counts.items()):
        flag = " \u26a0\ufe0f" if ct not in ("SUBSCRIPTION", "UNKNOWN") else ""
        ct_lines.append(f"<b>{ct}</b>: {cnt}{flag}")

    sections = [
        {"widgets": [{"textParagraph": {"text": (
            f"<b>{total}</b> active subscription projects  \u2022  "
            f"<b>{with_budget}</b> with budget  \u2022  "
            f"<b>{len(needs_fix)}</b> need contract fix"
        )}}]},
        {"header": "Contract Types",
         "widgets": [{"textParagraph": {"text": "<br>".join(ct_lines)}}]},
    ]

    if needs_fix:
        fix_lines = []
        for r in needs_fix:
            extra = ""
            if r["tm_budget_usd"]:
                extra = f" (${r['tm_budget_usd']:,.0f})"
            fix_lines.append(
                f"\u26a0\ufe0f [{r['contract_type']}] "
                f'<a href="{RL_APP_BASE}/{r["project_id"]}">{r["customer"]}</a>'
                f" {DASH} {r['project_name'][:50]}{extra}"
            )
        sections.append({
            "header": f"Correction Needed ({len(needs_fix)})",
            "widgets": [{"textParagraph": {"text": "<br>".join(fix_lines)}}],
        })

    card = {"cardsV2": [{"cardId": "subscription-audit", "card": {
        "header": {"title": f"Subscription Projects Audit {DASH} {date_str}",
                   "subtitle": "Post-Implementation Contract Review"},
        "sections": sections,
    }}]}

    if dry_run:
        print("  [DRY RUN] Would post audit summary to Google Chat")
        return

    data = json.dumps(card).encode()
    req = urllib.request.Request(
        GCHAT_WEBHOOK_URL, data=data,
        headers={"Content-Type": "application/json"}
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        if resp.status == 200:
            print("  Chat summary posted.")
        else:
            print(f"  WARNING: Chat post HTTP {resp.status}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Subscription Projects Audit")
    parser.add_argument("--dry-run", action="store_true", help="Preview without sending")
    args = parser.parse_args()

    if not API_KEY:
        print("ERROR: ROCKETLANE_API_KEY not set"); sys.exit(1)
    if not args.dry_run and (not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD):
        print("ERROR: GMAIL_ADDRESS and GMAIL_APP_PASSWORD required"); sys.exit(1)

    print("=" * 60)
    print("Subscription Projects Audit")
    print(f"Date: {NOW.strftime('%Y-%m-%d')} | Dry run: {args.dry_run}")
    print("=" * 60)

    # Fetch all projects
    print("Fetching all projects...")
    all_projects = fetch_all_projects()
    print(f"  {len(all_projects)} total")

    subs = [p for p in all_projects if is_post_impl(p) and is_active_subscription(p)]
    print(f"  {len(subs)} active subscription projects")

    # Fetch details in parallel
    print(f"Fetching details for {len(subs)} projects...")
    t_start = time.time()
    details = {}

    def _fetch(p):
        pid = p.get("projectId", "")
        try:
            return pid, fetch_project_detail(pid)
        except Exception as e:
            print(f"  WARN: {pid}: {e}")
            return pid, p

    with ThreadPoolExecutor(max_workers=4) as pool:
        futs = {pool.submit(_fetch, p): p for p in subs}
        done = 0
        for f in as_completed(futs):
            done += 1
            if done % 10 == 0:
                print(f"  {done}/{len(subs)}...")
            pid, d = f.result()
            details[pid] = d

    elapsed = time.time() - t_start
    print(f"  Completed in {elapsed:.0f}s")

    # Build rows
    rows = []
    for p in subs:
        pid = p.get("projectId", "")
        rows.append(extract_project_row(p, details.get(pid, p)))

    rows.sort(key=lambda r: (0 if r["needs_fix"] else 1, r["contract_type"], r["customer"]))

    needs_fix = sum(1 for r in rows if r["needs_fix"])
    with_budget = sum(1 for r in rows if r["total_budget_hrs"] > 0)
    print(f"\n  {len(rows)} projects | {with_budget} with budget | {needs_fix} need correction")

    # Build workbook
    print("\nBuilding spreadsheet...")
    wb = build_workbook(rows)

    # Email with attachment
    print("Sending audit email...")
    send_audit_email(wb, rows, dry_run=args.dry_run)

    # Chat summary
    print("Posting chat summary...")
    post_chat_summary(rows, dry_run=args.dry_run)

    print(f"\nDone. {len(rows)} projects audited.")


if __name__ == "__main__":
    main()
