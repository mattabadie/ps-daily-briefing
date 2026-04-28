#!/usr/bin/env python3
"""
Subscription Expansion Audit — LEGACY (slow, kept for fallback / workbook builder reuse)

Original implementation: fetch all 828 projects, filter client-side, per-project detail call,
per-project time-entry loop. ~239 API calls, ~3 min runtime.

Superseded by subscription_expansion_audit.py which uses native server-side filtering and
project-level rollup fields. This file is retained because it owns build_expansion_workbook(),
which the canonical script imports.

Output: Subscription_Expansion_Audit_legacy_<date>.xlsx
"""

import argparse
import os
import re
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

# ─── Load secrets ──────────────────────────────────────────────────────────────
SECRETS_FILE = Path("/Users/matthew.abadie/Library/Mobile Documents/com~apple~CloudDocs/iCloud Storage/Exterro/.secrets/rocketlane.env")
if SECRETS_FILE.exists():
    for line in SECRETS_FILE.read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            if v.strip():
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

# Reuse logic from existing scripts
from subscription_audit import (
    fetch_all_projects,
    fetch_project_detail,
    get_field,
    is_active_subscription,
    extract_project_row,
)
from subscription_tracker import (
    fetch_time_entries_for_project,
    compute_consumption,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NOW = datetime.now()


def enrich_row(p, detail, base_row, fetch_consumption=True):
    """Add expansion-analysis fields to a base row from extract_project_row()."""
    pid = p.get("projectId")

    # Modules + product family
    base_row["modules"] = get_field(p, "Opp: Product Names") or ""
    base_row["product_family"] = get_field(p, "Opp: Product Families") or ""

    # ARR — top-level annualizedRecurringRevenue, fall back to Opp field
    arr_top = detail.get("annualizedRecurringRevenue") or p.get("annualizedRecurringRevenue") or 0
    arr_opp = get_field(p, "Opp: Opportunity ARR") or 0
    try:
        arr_opp = float(arr_opp) if arr_opp else 0
    except (TypeError, ValueError):
        arr_opp = 0
    base_row["arr"] = arr_top or arr_opp or 0
    base_row["total_contract_value"] = get_field(p, "Opp: Total Contract Value") or ""

    # Customer contact + ownership
    base_row["account_owner"] = get_field(p, "Opp: Account Owner") or ""
    base_row["primary_contact"] = get_field(p, "Opp: Primary Customer Contact Name") or ""
    base_row["primary_contact_email"] = get_field(p, "Opp: Primary Customer Contact Email") or ""
    base_row["responsible_director"] = get_field(p, "Responsible Director") or ""

    # Subscription end date (renewal proxy)
    base_row["subscription_end"] = get_field(p, "Opp: Opportunity End Date") or ""

    # Health notes (free-text intel)
    health_notes = get_field(p, "Internal Project Health Notes") or ""
    base_row["health_notes"] = re.sub(r"<[^>]+>", "", str(health_notes))[:500].strip()

    # Consumption metrics
    base_row["hours_used"] = 0
    base_row["pct_consumed"] = 0
    base_row["last_entry_date"] = ""

    if fetch_consumption:
        try:
            entries = fetch_time_entries_for_project(pid)
            total_min = sum(e.get("minutes", 0) or 0 for e in entries)
            hrs_used = total_min / 60
            base_row["hours_used"] = round(hrs_used, 1)
            budgeted = base_row.get("total_budget_hrs", 0) or 0
            base_row["pct_consumed"] = round((hrs_used / budgeted * 100), 1) if budgeted > 0 else 0

            dates = []
            for e in entries:
                ds = e.get("date", "")
                if ds:
                    try:
                        dates.append(datetime.strptime(ds, "%Y-%m-%d"))
                    except ValueError:
                        pass
            base_row["last_entry_date"] = max(dates).strftime("%Y-%m-%d") if dates else ""
        except Exception as e:
            base_row["consumption_error"] = str(e)[:80]

    # Expansion-opportunity flags
    flags = []
    pct = base_row["pct_consumed"]
    if pct >= 100:
        flags.append("OVER budget")
    elif pct >= 75:
        flags.append("Renewal window")
    elif pct < 25 and base_row.get("total_budget_hrs", 0) > 0:
        flags.append("Underutilized")

    if not base_row["modules"]:
        flags.append("Module data missing")

    health = (base_row.get("health") or "").lower()
    if "red" in health:
        flags.append("Red health")
    elif "yellow" in health:
        flags.append("Yellow health")

    if base_row.get("needs_fix"):
        flags.append("Contract type wrong")

    seg = (base_row.get("client_segment") or "").lower()
    if "pinnacle" in seg or "strategic" in seg:
        flags.append(f"Tier: {base_row['client_segment']}")

    base_row["expansion_flags"] = " | ".join(flags) if flags else ""

    return base_row


def build_expansion_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscription Projects"

    headers = [
        "Customer", "Project Name", "SF Opportunity", "PSR",          # 1-4 (links up front)
        "Project ID", "Status", "PM",                                  # 5-7
        "Responsible Director", "Account Owner", "Primary Contact", "Contact Email",  # 8-11
        "Modules Licensed", "Product Family", "Client Segment",        # 12-14
        "ARR ($)", "Total Contract Value",                             # 15-16
        "Contract Type", "Frequency", "Periods", "Period Hrs", "Total Budget Hrs",  # 17-21
        "Hours Used", "% Consumed", "Last Entry",                      # 22-24
        "Period Budget ($)", "T&M Budget ($)",                         # 25-26
        "Subscription Start", "Subscription End", "Health",            # 27-29
        "Expansion Flags", "Health Notes (truncated)",                 # 30-31
    ]

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    expand_fill = PatternFill("solid", fgColor="DBEAFE")  # blue tint for expansion flag highlight
    data_font = Font(name="Arial", size=10)
    money_fmt = '$#,##0'
    pct_fmt = '0.0"%"'

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, 2):
        vals = [
            r["customer"],                          # 1 — Customer (hyperlinked to SF Account)
            r["project_name"],                      # 2
            r["opp_url"],                           # 3 — SF Opportunity
            r.get("psr_url", ""),                   # 4 — PSR
            r["project_id"],                        # 5
            r["status"],                            # 6
            r["pm"],                                # 7
            r.get("responsible_director", ""),      # 8
            r.get("account_owner", ""),             # 9
            r.get("primary_contact", ""),           # 10
            r.get("primary_contact_email", ""),     # 11
            r.get("modules", ""),                   # 12
            r.get("product_family", ""),            # 13
            r.get("client_segment", ""),            # 14
            r.get("arr", 0) or 0,                   # 15 — ARR ($)
            r.get("total_contract_value", ""),      # 16
            r["contract_type"],                     # 17
            r["frequency"],                         # 18
            r["periods"],                           # 19
            r["period_hrs"],                        # 20
            r["total_budget_hrs"],                  # 21
            r.get("hours_used", 0),                 # 22
            r.get("pct_consumed", 0),               # 23 — % Consumed
            r.get("last_entry_date", ""),           # 24
            r["period_budget_usd"],                 # 25 — $
            r["tm_budget_usd"],                     # 26 — $
            r["start_date"],                        # 27
            r.get("subscription_end", ""),          # 28
            r["health"],                            # 29
            r.get("expansion_flags", ""),           # 30
            r.get("health_notes", ""),              # 31
        ]
        link_font = Font(name="Arial", size=10, color="3B82F6", underline="single")
        psr_missing_fill = PatternFill("solid", fgColor="FCA5A5")
        customer_link_font = Font(name="Arial", size=10, color="1E3A8A", underline="single", bold=True)
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = data_font
            c.border = thin_border
            # Customer (col 1) — hyperlink to SF Account if we have account URL
            if col == 1:
                acct_url = r.get("sf_account_url")
                if acct_url:
                    c.hyperlink = acct_url
                    c.font = customer_link_font
            # SF Opportunity hyperlink (col 3, was 30)
            if col == 3:
                if v:
                    c.value = "View Opp"
                    c.hyperlink = v
                    c.font = link_font
                else:
                    c.value = ""
            # PSR hyperlink + missing-flag (col 4, was 31)
            if col == 4:
                if v:
                    psr_id = (r.get("psr_external_id") or "")[:6]
                    c.value = f"View PSR {psr_id}" if psr_id else "View PSR"
                    c.hyperlink = v
                    c.font = link_font
                else:
                    c.value = "MISSING"
                    c.fill = psr_missing_fill
                    c.font = Font(name="Arial", size=10, bold=True, color="991B1B")
            # money columns: ARR (15), Period Budget (25), T&M Budget (26)
            if col in (15, 25, 26):
                c.number_format = money_fmt
            # percent column (23)
            if col == 23:
                c.number_format = pct_fmt
            # contract type wrong → red tint on contract type cell (17)
            if col == 17 and r.get("needs_fix"):
                c.fill = red_fill
            # health color code (29)
            if col == 29:
                hl = (v or "").lower()
                if "red" in hl:
                    c.fill = PatternFill("solid", fgColor="FCA5A5")
                elif "yellow" in hl:
                    c.fill = yellow_fill
                elif "green" in hl:
                    c.fill = green_fill
            # expansion flags highlight if any flag set (30)
            if col == 30 and v:
                c.fill = expand_fill

    widths = [
        28, 50, 14, 16,                    # 1-4 (customer / project name / opp link / psr link)
        12, 14, 20,                        # 5-7 (project id / status / pm)
        24, 18, 22, 28,                    # 8-11 (director/owner/contact/email)
        30, 18, 14,                        # 12-14 (modules/family/segment)
        14, 16,                            # 15-16 (ARR/TCV)
        18, 12, 8, 10, 14,                 # 17-21 (contract type/frequency/budget)
        12, 12, 12,                        # 22-24 (used/pct/last entry)
        14, 14,                            # 25-26 ($)
        14, 14, 10,                        # 27-29 (start/end/health)
        38, 50,                            # 30-31 (flags/notes)
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Expansion Summary sheet
    ws2 = wb.create_sheet("Expansion Summary", 0)  # put first
    ws2.sheet_properties.tabColor = "3B82F6"

    ws2["A1"] = "Subscription Expansion Snapshot"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1E293B")
    ws2["A2"] = f"Generated {NOW.strftime('%Y-%m-%d %H:%M')} | All active subscription projects across all teams"
    ws2["A2"].font = Font(name="Arial", italic=True, size=10, color="475569")

    # Top stats
    bold = Font(name="Arial", bold=True, size=11)
    norm = Font(name="Arial", size=10)
    total = len(rows)
    with_arr = sum(1 for r in rows if (r.get("arr") or 0) > 0)
    arr_total = sum((r.get("arr") or 0) for r in rows)
    over_budget = sum(1 for r in rows if (r.get("pct_consumed") or 0) >= 100)
    in_renewal = sum(1 for r in rows if 75 <= (r.get("pct_consumed") or 0) < 100)
    underutilized = sum(1 for r in rows if 0 < (r.get("pct_consumed") or 0) < 25 and r.get("total_budget_hrs", 0) > 0)
    needs_fix = sum(1 for r in rows if r.get("needs_fix"))
    no_modules = sum(1 for r in rows if not r.get("modules"))
    no_psr = sum(1 for r in rows if not r.get("psr_url"))

    stats = [
        ("Total active subscription projects", total),
        ("Total ARR across portfolio", f"${arr_total:,.0f}"),
        ("Projects with ARR populated", f"{with_arr} ({with_arr*100//total if total else 0}%)"),
        ("OVER budget (≥100% consumed) — expand or renew NOW", over_budget),
        ("In renewal window (75-99% consumed)", in_renewal),
        ("Underutilized (<25% consumed) — service-led activation play", underutilized),
        ("Contract type misconfigured — fix before reporting", needs_fix),
        ("Missing module data — sales hygiene gap", no_modules),
        ("Missing PSR link — broken SFDC↔Rocketlane integration", no_psr),
    ]
    for i, (label, val) in enumerate(stats, 4):
        ws2.cell(row=i, column=1, value=label).font = bold
        c = ws2.cell(row=i, column=2, value=val)
        c.font = norm
        if "OVER" in label or "renewal window" in label:
            c.fill = expand_fill
        if "misconfigured" in label or "Missing" in label:
            c.fill = red_fill if int(val) > 0 else c.fill

    # Top 10 customers by ARR
    ws2["A15"] = "Top 10 Customers by ARR"
    ws2["A15"].font = Font(name="Arial", bold=True, size=12)
    ws2["A16"] = "Customer"
    ws2["B16"] = "ARR"
    ws2["C16"] = "Modules"
    for h in ("A16", "B16", "C16"):
        ws2[h].font = bold

    customer_arr = {}
    customer_modules = {}
    for r in rows:
        c = r["customer"]
        customer_arr[c] = customer_arr.get(c, 0) + (r.get("arr") or 0)
        customer_modules.setdefault(c, set()).update(
            m.strip() for m in (r.get("modules") or "").split(",") if m.strip()
        )
    top10 = sorted(customer_arr.items(), key=lambda x: x[1], reverse=True)[:10]
    for i, (cust, arr) in enumerate(top10, 17):
        ws2.cell(row=i, column=1, value=cust).font = norm
        c2 = ws2.cell(row=i, column=2, value=arr)
        c2.font = norm
        c2.number_format = money_fmt
        ws2.cell(row=i, column=3, value=", ".join(sorted(customer_modules.get(cust, set())))).font = norm

    # Module distribution
    off = 29
    ws2.cell(row=off, column=1, value="Module / Product Family Distribution").font = Font(name="Arial", bold=True, size=12)
    family_counts = Counter(r.get("product_family", "Unknown") or "Unknown" for r in rows)
    ws2.cell(row=off + 2, column=1, value="Product Family").font = bold
    ws2.cell(row=off + 2, column=2, value="Project Count").font = bold
    for i, (fam, cnt) in enumerate(family_counts.most_common(), off + 3):
        ws2.cell(row=i, column=1, value=fam).font = norm
        ws2.cell(row=i, column=2, value=cnt).font = norm

    ws2.column_dimensions["A"].width = 56
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 60

    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--quick", action="store_true",
                        help="Skip per-project time-entry calls (fast, no %% consumed)")
    parser.add_argument("--workspace", default=None,
                        help="Output folder. Default: this script's folder")
    args = parser.parse_args()

    if not os.environ.get("ROCKETLANE_API_KEY"):
        print("ERROR: ROCKETLANE_API_KEY not set. Check .secrets/rocketlane.env.", file=sys.stderr)
        sys.exit(1)

    print("=" * 70)
    print("Subscription Expansion Audit (legacy)")
    print(f"Date: {NOW.strftime('%Y-%m-%d %H:%M')} | Quick mode: {args.quick}")
    print("=" * 70)

    print("Fetching all projects...")
    all_projects = fetch_all_projects()
    print(f"  {len(all_projects)} total projects in Rocketlane")

    subs = [p for p in all_projects if is_active_subscription(p)]
    print(f"  {len(subs)} active subscription projects (across ALL teams)")

    print(f"Fetching project detail{'+ time entries' if not args.quick else ''} for {len(subs)} projects...")
    rows = []
    t0 = time.time()

    def process_one(p):
        pid = p["projectId"]
        try:
            detail = fetch_project_detail(pid)
        except Exception:
            detail = p
        base = extract_project_row(p, detail)
        return enrich_row(p, detail, base, fetch_consumption=not args.quick)

    # Parallelize lightly for speed
    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {pool.submit(process_one, p): p for p in subs}
        for i, f in enumerate(as_completed(futures), 1):
            try:
                rows.append(f.result())
            except Exception as e:
                p = futures[f]
                print(f"  WARN {p.get('projectName','?')[:40]}: {e}")
            if i % 10 == 0 or i == len(subs):
                print(f"  {i}/{len(subs)} ({time.time()-t0:.0f}s elapsed)")

    rows.sort(key=lambda r: (r.get("arr") or 0), reverse=True)

    # Stats
    print()
    print("=" * 70)
    print("RESULTS")
    print("=" * 70)
    arr_total = sum((r.get("arr") or 0) for r in rows)
    print(f"Total active subscription projects: {len(rows)}")
    print(f"Total ARR: ${arr_total:,.0f}")
    if not args.quick:
        over = sum(1 for r in rows if (r.get("pct_consumed") or 0) >= 100)
        renew = sum(1 for r in rows if 75 <= (r.get("pct_consumed") or 0) < 100)
        under = sum(1 for r in rows if 0 < (r.get("pct_consumed") or 0) < 25 and r.get("total_budget_hrs", 0) > 0)
        print(f"  OVER budget: {over}  |  In renewal window: {renew}  |  Underutilized: {under}")
    needs_fix = sum(1 for r in rows if r.get("needs_fix"))
    no_modules = sum(1 for r in rows if not r.get("modules"))
    print(f"  Contract type wrong: {needs_fix}  |  Missing modules: {no_modules}")

    print("\nBuilding spreadsheet...")
    wb = build_expansion_workbook(rows)
    outputs_dir = args.workspace or "/Users/matthew.abadie/Library/Mobile Documents/com~apple~CloudDocs/iCloud Storage/Exterro/outputs"
    os.makedirs(outputs_dir, exist_ok=True)
    fname = f"Subscription_Expansion_Audit_legacy_{NOW.strftime('%Y%m%d')}.xlsx"
    out_path = os.path.join(outputs_dir, fname)
    wb.save(out_path)
    print(f"  Saved: {out_path}")


if __name__ == "__main__":
    main()
