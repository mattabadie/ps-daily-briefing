"""
z2e_migration_audit.py — Z2E (Zapproved-to-Exterro) migration program audit.

Pulls the Z2E cohort across all three sub-type values, milestones, and time
entries, then writes a multi-tab Excel workbook with health, ramp velocity,
and capacity / burn analysis.

Z2E commercial model:
  Z2E migration projects are delivered NO-CHARGE as a customer-retention play
  post-Exterro's acquisition of Zapproved. Hours are tracked for capacity
  purposes — these projects are NOT revenue-generating. The relevant revenue
  lens is ARR retention / churn risk, not project revenue.

Field nuance (verified 2026-04-28):
  - The custom field is exactly  `eDisc: Project Sub-Type`  (hyphenated).
  - It is multi-value / array: a project can be tagged ["Z2E", "Add On"].
  - It is sometimes missing from search-response payloads — get-by-id returns
    it reliably; the bulk projects endpoint with includeAllFields=true also
    returns it.
  - The MCP wrapper's customFieldFiltersJson rejects this field. Native API
    `project.field.<id>.oneOf=...` works. This script uses the native API.

Usage:
    python z2e_migration_audit.py [--since 2025-10-01]
                                  [--arr-csv /path/to/customers_export.csv]
                                  [--out Z2E_Audit_2026-04-28.xlsx]

Env / secrets:
    ROCKETLANE_API_KEY auto-loaded by rocketlane_client from
    Exterro/.secrets/rocketlane.env.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import urllib.parse
from collections import defaultdict
from datetime import datetime, timedelta, date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from rocketlane_client import (
    api_get,
    fetch_bulk_time_entries,
    fetch_time_entries_per_project,
    group_entries_by_project,
)


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
SUBTYPE_FIELD_LABEL = "eDisc: Project Sub-Type"
# Ordered MOST SPECIFIC FIRST. Projects can be multi-tagged (e.g.,
# ["Z2E", "Z2E Phase 1", "Add On"]) — when checking bucket membership,
# Phase 1 / Not Started win over the generic "Z2E" tag.
Z2E_VALUES = ["Z2E Phase 1", "Z2E - Not Started", "Z2E"]

# Bucket label rendered in the workbook
PHASE_LABEL = {
    "Z2E Phase 1":      "Phase 1 — Tech Integration",
    "Z2E":              "Phase 2 — Migration & Ramp",
    "Z2E - Not Started": "Not Started",
}

# Default time-entry window (used by burn analysis)
DEFAULT_SINCE = "2025-10-01"

# Standard outputs folder for shareable artifacts (Excel, CSV, JSON).
# All ps-daily-briefing scripts write here unless --out is overridden.
DEFAULT_OUTPUT_DIR = Path(
    "/Users/matthew.abadie/Library/Mobile Documents/com~apple~CloudDocs/"
    "iCloud Storage/Exterro/Outputs"
)

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM FIELD IDs (verified against Travelers / TrueBlue payloads 2026-04-28)
# Field IDs are stable; field labels can drift. Prefer IDs.
# ─────────────────────────────────────────────────────────────────────────────
F_RAG_HEALTH        = 2050224  # Red/Yellow/Green Health (single-choice: Green/Yellow/Red)
F_INTERNAL_STATUS   = 2050226  # Internal Weekly Status (HTML, dated narrative entries)
F_HEALTH_NOTES      = 1649519  # Internal Project Health Notes (short summary)
F_SUBTYPE           = 2046191  # eDisc: Project Sub-Type (multi-choice)
F_GO_LIVE_PLANNED   = 2039402  # eDisc: Go Live - Planned (date)
F_KICKOFF_ACTUAL    = 2039419  # eDisc: Time to Kickoff (days)
F_KICKOFF_PLANNED   = 2039420  # eDisc: Time to Kickoff - Planned (days)
F_TTV_PLANNED       = 2039412  # eDisc: TTV - Planned (days)
F_IMPL_CYCLE        = 2039423  # eDisc: Implementation Cycle Time (days)
F_TOTAL_DURATION    = 2039428  # eDisc: Total Project Duration - Planned (days)
F_MODULES           = 2051607  # eDisc: Modules to Implement (multi-choice)
F_INTEGRATIONS      = 2051608  # eDisc: Integrations in Scope (multi-choice)
F_TENANT_DOMAIN     = 1854718  # Tenant Domain (text)
F_OPP_TCV           = 2016832  # Opp: Total Contract Value (number)
F_OPP_ACCOUNT_OWNER = 2094159  # Opp: Account Owner (text)
F_OPP_OWNER         = 1875409  # Opp: Opportunity Owner (text)
F_OPP_TYPE          = 1875408  # Opp: Opportunity Type (single-choice)
F_OPP_URL           = 1875410  # Opp: Opportunity URL (text)
F_CLIENT_SEG        = 2036126  # Client Segmentation (single-choice: Strategic/Growth/etc.)
F_FORECAST_CONF     = 2106520  # PSR: Forecast Confidence (single-choice)

# Cutoffs for Go-Live Cadence tracking. Edit list to change reporting buckets.
GO_LIVE_CUTOFFS = [
    ("Through 6/30/2026", date(2026, 6, 30)),
    ("Through 9/30/2026", date(2026, 9, 30)),
    ("Through 12/31/2026", date(2026, 12, 31)),
]

# Status values that indicate the project has actually gone live (migrated).
# Used as a fallback when the Go-Lives (2) tab has no record for a project.
# The Go-Lives (2) tab is the primary source of truth for go-live dates.
LIVE_STATUSES = {"hypercare", "completed", "closeout", "partially live"}

# Status values that mean the project is DONE — for the --remaining-only view.
# Cancelled = churned. Completed/Closeout = wrapped. Hypercare/Partially Live =
# live in production, just stabilizing → treated as done for the remaining-work
# focus list. Edit as needed.
COMPLETED_STATUSES = {
    "completed", "closeout", "cancelled",
    "hypercare", "partially live",
}

# Forward-looking analysis cutoff. Anything with an Actual go-live before this
# date is excluded from the cohort entirely — already migrated, no longer
# part of the program-tracking metric.
FORWARD_LOOKING_CUTOFF = date(2026, 1, 1)

# Phase Bucket values that DO count toward the 6/30, 9/30, 12/31 ARR
# milestone totals. Phase 1 (Tech Integration) is excluded — each Phase 1
# project has a companion Phase 2 project that is the actual migration.
MILESTONE_PHASES = {"Z2E", "Z2E - Not Started"}

# Snapshot directory — each run drops a JSON snapshot for trend tracking.
SNAPSHOT_DIR = DEFAULT_OUTPUT_DIR / "z2e_snapshots"

# Separate snapshot directory for the --remaining-only mode so the trend
# remains internally consistent (its cohort changes as projects complete).
REMAINING_SNAPSHOT_DIR = DEFAULT_OUTPUT_DIR / "z2e_remaining_snapshots"

# Excel styling
HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER  = Border(*[Side(style="thin", color="BFBFBF") for _ in range(4)])
HEALTH_RED    = PatternFill("solid", fgColor="F8CBAD")
HEALTH_YELLOW = PatternFill("solid", fgColor="FFE699")
HEALTH_GREEN  = PatternFill("solid", fgColor="C6EFCE")


# ─────────────────────────────────────────────────────────────────────────────
# ROCKETLANE FETCHERS
# ─────────────────────────────────────────────────────────────────────────────
def find_field_id(label: str) -> int | None:
    """Walk /fields pages until a field with this label is found.

    Rocketlane's /fields endpoint is paginated (page size 100). The Z2E sub-type
    field is past page 1 of an org with hundreds of custom fields, so this must
    paginate.
    """
    token = None
    while True:
        url = "fields?pageSize=100"
        if token:
            url += f"&pageToken={urllib.parse.quote(token)}"
        resp = api_get(url)
        for f in resp.get("data", []):
            if f.get("fieldLabel") == label:
                return f.get("fieldId")
        pag = resp.get("pagination", {})
        if pag.get("hasMore") and pag.get("nextPageToken"):
            token = pag["nextPageToken"]
        else:
            return None


def fetch_z2e_projects(field_id: int) -> tuple[list[dict], int]:
    """Fetch Z2E project cohort using native server-side `projectName.cn`.

    The custom-field filter on `eDisc: Project Sub-Type` (fieldId 2046191)
    returns 500 INTERNAL_SERVER_ERROR for both `.value=` and `.oneOf=` operators
    (Rocketlane backend bug, verified 2026-04-28). The MCP wrapper's
    `customFieldFiltersJson` likewise rejects it.

    Native `projectName.cn=` (contains) is the documented filter on the same
    /projects endpoint and works fine. Z2E projects almost always have "Z2E"
    or "Zapproved" in the name — two calls catch the cohort with high recall.

    A handful of edge-case projects (sub-type tagged but named differently)
    will be missed by name search alone. Those will surface in any cross-check
    against subscription analyses and can be added by name if found.
    """
    cutoff = datetime(2024, 1, 1)
    cutoff_ms = int(cutoff.timestamp() * 1000)

    seen: dict[int, dict] = {}
    total_calls = 0

    for needle in ("Z2E", "Zapproved"):
        qs = (
            f"projectName.cn={urllib.parse.quote(needle, safe='')}"
            f"&updatedAt.ge={cutoff_ms}"
            f"&includeAllFields=true"
            f"&pageSize=100"
        )
        try:
            projects, calls = _paginate_projects(qs)
        except Exception as e:
            print(f"      WARN projectName.cn={needle} failed: {str(e)[:120]}")
            continue
        total_calls += calls
        for p in projects:
            pid = p.get("projectId")
            if pid is not None:
                seen[pid] = p
        print(f"      projectName.cn='{needle}' → {len(projects)} projects "
              f"({calls} call(s))")

    return list(seen.values()), total_calls


def _paginate_projects(qs: str) -> tuple[list[dict], int]:
    all_p, token, calls = [], None, 0
    while True:
        url = f"projects?{qs}"
        if token:
            url += f"&pageToken={urllib.parse.quote(token)}"
        resp = api_get(url)
        calls += 1
        all_p.extend(resp.get("data", []))
        pag = resp.get("pagination", {})
        if pag.get("hasMore") and pag.get("nextPageToken"):
            token = pag["nextPageToken"]
        else:
            break
    return all_p, calls


def _get_subtype_tags(project: dict) -> list[str]:
    """Return ALL tag values from the eDisc: Project Sub-Type field on a
    project, as a list. Empty list if field absent."""
    for f in project.get("fields") or []:
        if f.get("fieldLabel") != SUBTYPE_FIELD_LABEL:
            continue
        raw = f.get("fieldValueLabel")
        if raw is None:
            raw = f.get("fieldValue")
        if raw is None:
            return []
        if isinstance(raw, list):
            return [str(v).strip() for v in raw if str(v).strip()]
        if isinstance(raw, str):
            return [v.strip() for v in raw.split(",") if v.strip()]
        return [str(raw)]
    return []


def get_subtype_value(project: dict) -> str | None:
    """Return the bucketing key for a project — the most-specific Z2E sub-type
    tag found. Z2E_VALUES is ordered specific-first so a project tagged with
    BOTH ["Z2E", "Z2E Phase 1"] returns "Z2E Phase 1" (Phase 1 bucket)."""
    tags = _get_subtype_tags(project)
    for preferred in Z2E_VALUES:
        if preferred in tags:
            return preferred
    return None


def get_all_subtype_tags(project: dict) -> str:
    """Comma-joined string of all tags on the sub-type field, for display."""
    return ", ".join(_get_subtype_tags(project))


# ─────────────────────────────────────────────────────────────────────────────
# Generic field-by-id accessor + typed helpers
# ─────────────────────────────────────────────────────────────────────────────
def field_by_id(project: dict, field_id: int) -> dict | None:
    """Return the {fieldId, fieldValue, fieldValueLabel} dict for the given
    custom-field id on a project, or None."""
    for f in project.get("fields") or []:
        if f.get("fieldId") == field_id:
            return f
    return None


def field_label(project: dict, field_id: int) -> str | None:
    """Return the human-readable label/value for a custom field by id."""
    f = field_by_id(project, field_id)
    if not f:
        return None
    val = f.get("fieldValueLabel")
    if val is None:
        val = f.get("fieldValue")
    if val is None:
        return None
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val)


def field_number(project: dict, field_id: int) -> float | None:
    """Return a numeric custom field value (e.g., ARR, TCV, days)."""
    f = field_by_id(project, field_id)
    if not f:
        return None
    raw = f.get("fieldValue")
    if raw is None:
        raw = f.get("fieldValueLabel")
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def field_date(project: dict, field_id: int) -> date | None:
    """Return a date-typed custom field as a python date."""
    f = field_by_id(project, field_id)
    if not f:
        return None
    raw = f.get("fieldValue") or f.get("fieldValueLabel")
    if not raw:
        return None
    s = str(raw)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def field_count(project: dict, field_id: int) -> int:
    """Count items in a multi-select / multi-value custom field."""
    f = field_by_id(project, field_id)
    if not f:
        return 0
    raw = f.get("fieldValue")
    if isinstance(raw, list):
        return len(raw)
    label = f.get("fieldValueLabel")
    if isinstance(label, str) and label:
        return len([x for x in label.split(",") if x.strip()])
    return 0


# RAG normaliser — field returns "Green"/"Yellow"/"Red" labels
def get_rag(project: dict) -> str | None:
    val = field_label(project, F_RAG_HEALTH)
    if not val:
        return None
    s = val.strip().title()
    if s in ("Green", "Yellow", "Red"):
        return s
    return s


# Latest weekly status entry — extract first paragraph from the dated HTML log
_HTML_TAG_RE = re.compile(r"<[^>]+>")
_DATE_LEADER_RE = re.compile(r"^\s*(\d{1,2}/\d{1,2}(?:/\d{2,4})?(?:\s+\w+)?:?)\s*", re.UNICODE)


def get_latest_weekly_status(project: dict) -> str | None:
    """Extract the most-recent dated entry from `Internal Weekly Status`.
    The field is HTML with bold-tagged dated entries separated by <br><br>."""
    f = field_by_id(project, F_INTERNAL_STATUS)
    if not f:
        return None
    raw = f.get("fieldValueLabel") or f.get("fieldValue") or ""
    if not raw:
        return None
    # Split into entries on double-br, then strip HTML
    parts = re.split(r"(?:<br\s*/?>\s*){2,}", str(raw))
    if not parts:
        return None
    first = _HTML_TAG_RE.sub(" ", parts[0])
    first = re.sub(r"\s+", " ", first).strip()
    return first[:500] if first else None


# Need `re` — add to imports section if missing


def fetch_milestones(project_id: int) -> list[dict]:
    """Per-project milestones.

    NOTE: Both `/projects/{id}/milestones` and `/projects/{id}/phases` return
    404 against `services.api.exterro.com/api/1.0` (verified 2026-04-28). The
    MCP wrappers `rocketlane_get_project_milestones` and `_get_project_phases`
    exist, so the underlying URL must be elsewhere — possibly under v1, or
    routed through a different path. Until that's nailed down, return [] and
    let the workbook ship without milestone analysis.
    """
    return []


# ─────────────────────────────────────────────────────────────────────────────
# ARR JOIN — from CSV or XLSX export of the Z2E tracking spreadsheet
# ─────────────────────────────────────────────────────────────────────────────
def load_arr_file(path: str) -> dict[str, float]:
    """Dispatch loader by file extension. Accepts .csv or .xlsx (.xlsm)."""
    p = Path(path)
    if not p.exists():
        print(f"  WARN: ARR file not found at {path}")
        return {}
    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _load_arr_xlsx(p)
    return _load_arr_csv_internal(p)


def _load_arr_xlsx(p: Path) -> dict[str, float]:
    """Walk every sheet in the workbook, find the row containing
    'ARR from Finance' (case-insensitive) and a customer-name column,
    parse the data rows below."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  WARN: openpyxl not installed — cannot read xlsx ARR file")
        return {}

    wb = load_workbook(p, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx, name_col, arr_col = _locate_arr_columns(rows)
        if header_idx is None:
            continue
        arr = _parse_arr_rows(rows, header_idx, name_col, arr_col)
        if arr:
            print(f"  loaded ARR from sheet '{sheet_name}' "
                  f"(header at row {header_idx + 1}, "
                  f"name col={name_col + 1}, arr col={arr_col + 1})")
            return arr
    print(f"  WARN: 'ARR from Finance' header not found in any sheet of {p.name}")
    return {}


def _load_arr_csv_internal(p: Path) -> dict[str, float]:
    with p.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    header_idx, name_col, arr_col = _locate_arr_columns(rows)
    if header_idx is None:
        print(f"  WARN: 'ARR from Finance' header not found in {p.name}")
        return {}
    return _parse_arr_rows(rows, header_idx, name_col, arr_col)


def _locate_arr_columns(rows: list) -> tuple[int | None, int, int]:
    """Heuristic header row finder. Looks at first ~30 rows for a cell
    matching 'ARR from Finance' (case-insensitive). Returns
    (header_idx, name_col, arr_col) — or (None, 0, 0) if not found."""
    for i, row in enumerate(rows[:30]):
        if not row:
            continue
        arr_col = None
        for j, cell in enumerate(row):
            if cell and "arr from finance" in str(cell).lower():
                arr_col = j
                break
        if arr_col is None:
            continue
        # Customer-name column: look for "customer" / "company" / "account" in the header,
        # else default to first non-empty cell index
        name_col = 0
        for j, cell in enumerate(row):
            cs = str(cell or "").lower()
            if any(tok in cs for tok in ("customer", "company name", "account")):
                name_col = j
                break
        return i, name_col, arr_col
    return None, 0, 0


def _parse_arr_rows(rows: list, header_idx: int, name_col: int, arr_col: int) -> dict[str, float]:
    """Read data rows below the header, build {customer_lower: arr_float}."""
    arr: dict[str, float] = {}
    for row in rows[header_idx + 1:]:
        if row is None or len(row) <= max(name_col, arr_col):
            continue
        name = str(row[name_col] or "").strip()
        if not name:
            continue
        raw = str(row[arr_col] or "").strip().replace("$", "").replace(",", "").replace(" ", "")
        if not raw or raw.lower() in ("none", "n/a", "-"):
            continue
        try:
            arr[name.lower()] = float(raw)
        except ValueError:
            pass
    return arr


# Backwards-compat shim: existing code calls load_arr_csv()
def load_arr_csv(path: str) -> dict[str, float]:
    """Backwards-compat wrapper. Use load_arr_file directly going forward."""
    return load_arr_file(path)


# ─────────────────────────────────────────────────────────────────────────────
# GO-LIVES (2) — authoritative actual/estimated go-live dates per customer
# ─────────────────────────────────────────────────────────────────────────────
def load_go_lives_xlsx(path: str) -> dict[str, dict]:
    """Read the Go-Lives (2) tab from the Z2E tracking xlsx.

    Returns a map keyed by customer name (lowercased) AND by SF Account ID
    where available, so callers can resolve by either. Each value is:
        {
            "account_name":  str,
            "go_live_date":  date or None,
            "go_live_type":  "Actual" | "Estimated" | None,
            "status":        str (e.g., "Migration complete"),
            "current_arr":   float or None,
            "sf_id":         str (SF 18-digit ID) or None,
        }
    """
    p = Path(path)
    if not p.exists() or p.suffix.lower() not in (".xlsx", ".xlsm"):
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  WARN: openpyxl not installed — cannot read Go-Lives (2)")
        return {}

    target_sheet = None
    wb = load_workbook(p, data_only=True, read_only=True)
    for name in ("Go-Lives (2)", "Go Lives (2)", "GoLives (2)"):
        if name in wb.sheetnames:
            target_sheet = name
            break
    if target_sheet is None:
        print(f"  WARN: 'Go-Lives (2)' tab not found in {p.name}")
        return {}

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Heuristic header location — column index lookup based on label substrings
    header = rows[0]
    def col_idx(*tokens) -> int | None:
        for i, cell in enumerate(header):
            if not cell:
                continue
            cs = str(cell).lower()
            if any(t in cs for t in tokens):
                return i
        return None

    c_id     = col_idx("18 digit", "sf id", "salesforce id")
    c_name   = col_idx("account name", "customer", "company")
    c_arr    = col_idx("current arr", "arr")
    c_status = col_idx("project status", "status")
    c_date   = col_idx("est/act go live", "go live date", "go-live")
    c_type   = col_idx("estimated or actual", "est/actual", "estimate or actual")

    if c_name is None or c_date is None:
        print(f"  WARN: Go-Lives (2) header missing name or date column "
              f"(name={c_name}, date={c_date})")
        return {}

    out: dict[str, dict] = {}
    for row in rows[1:]:
        if not row:
            continue
        name = (row[c_name] if c_name < len(row) else None) or ""
        name = str(name).strip()
        if not name:
            continue
        d_raw = row[c_date] if c_date is not None and c_date < len(row) else None
        go_live_d = None
        if hasattr(d_raw, "year"):
            go_live_d = d_raw.date() if hasattr(d_raw, "date") else d_raw
        gl_type = None
        if c_type is not None and c_type < len(row):
            tv = row[c_type]
            gl_type = str(tv).strip() if tv else None
        status = ""
        if c_status is not None and c_status < len(row):
            sv = row[c_status]
            status = str(sv).strip() if sv else ""
        sf_id = None
        if c_id is not None and c_id < len(row):
            iv = row[c_id]
            sf_id = str(iv).strip() if iv else None
        cur_arr = None
        if c_arr is not None and c_arr < len(row):
            av = row[c_arr]
            try:
                cur_arr = float(av) if av not in (None, "") else None
            except (TypeError, ValueError):
                cur_arr = None

        record = {
            "account_name": name,
            "go_live_date": go_live_d,
            "go_live_type": gl_type,
            "status":       status,
            "current_arr":  cur_arr,
            "sf_id":        sf_id,
        }
        out[name.lower()] = record
        if sf_id:
            out[f"sfid:{sf_id}"] = record
    return out


def lookup_go_live(go_lives: dict, customer_name: str,
                   sf_id: str | None = None) -> dict | None:
    """Resolve a project to its Go-Lives (2) record by SF ID first, then name."""
    if not go_lives:
        return None
    if sf_id:
        rec = go_lives.get(f"sfid:{sf_id}")
        if rec:
            return rec
    if customer_name:
        return go_lives.get(customer_name.lower())
    return None


def is_excluded_pre_2026(go_lives_record: dict | None) -> bool:
    """True if the project went live before 2026-01-01 (Actual). These are
    excluded from the forward-looking analysis."""
    if not go_lives_record:
        return False
    if (go_lives_record.get("go_live_type") or "").lower() != "actual":
        return False
    d = go_lives_record.get("go_live_date")
    if not d:
        return False
    return d < FORWARD_LOOKING_CUTOFF


# ─────────────────────────────────────────────────────────────────────────────
# DERIVED METRICS
# ─────────────────────────────────────────────────────────────────────────────
def parse_iso_date(val) -> date | None:
    if not val:
        return None
    if isinstance(val, dict):
        val = val.get("value") or val.get("date")
    if not val:
        return None
    s = str(val)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def project_row(p: dict, ms_list: list[dict], te_list: list[dict],
                arr_lookup: dict[str, float],
                go_lives: dict[str, dict] | None = None) -> dict:
    """Compute the per-project derived row used in the workbook."""
    go_lives = go_lives or {}
    customer = (p.get("customer") or {}).get("companyName") or \
               (p.get("customer") or {}).get("name") or ""
    owner = ((p.get("owner") or {}).get("firstName", "") + " " +
             (p.get("owner") or {}).get("lastName", "")).strip()
    status = (p.get("status") or {}).get("label") or ""

    start_d = parse_iso_date(p.get("startDate"))
    due_d   = parse_iso_date(p.get("dueDate"))
    today   = datetime.now().date()

    days_in_phase = (today - start_d).days if start_d else None
    days_to_due   = (due_d - today).days if due_d else None

    # Project-level rollups (server-computed, on the project payload)
    tracked_h   = p.get("trackedHours") or (p.get("trackedMinutes") or 0) / 60

    # New: real RAG, contract metrics, schedule targets, complexity, ownership
    rag             = get_rag(p)
    tcv             = field_number(p, F_OPP_TCV)
    client_seg      = field_label(p, F_CLIENT_SEG)
    opp_type        = field_label(p, F_OPP_TYPE)
    forecast_conf   = field_label(p, F_FORECAST_CONF)
    account_owner   = field_label(p, F_OPP_ACCOUNT_OWNER)
    opp_owner       = field_label(p, F_OPP_OWNER)
    tenant_domain   = field_label(p, F_TENANT_DOMAIN)
    opp_url         = field_label(p, F_OPP_URL)
    go_live_planned = field_date(p, F_GO_LIVE_PLANNED)
    kickoff_actual  = field_number(p, F_KICKOFF_ACTUAL)
    kickoff_planned = field_number(p, F_KICKOFF_PLANNED)
    ttv_planned     = field_number(p, F_TTV_PLANNED)
    impl_cycle      = field_number(p, F_IMPL_CYCLE)
    total_duration  = field_number(p, F_TOTAL_DURATION)
    n_modules       = field_count(p, F_MODULES)
    n_integrations  = field_count(p, F_INTEGRATIONS)
    modules_list    = field_label(p, F_MODULES) or ""
    integrations_list = field_label(p, F_INTEGRATIONS) or ""
    weekly_status   = get_latest_weekly_status(p)

    kickoff_slip = None
    if kickoff_actual is not None and kickoff_planned is not None:
        kickoff_slip = kickoff_actual - kickoff_planned

    # Days-to-Go-Live (negative = past target)
    days_to_go_live = None
    if go_live_planned:
        days_to_go_live = (go_live_planned - today).days

    # ARR — only from the trusted Google Sheet CSV. Salesforce ARR is unreliable.
    arr = arr_lookup.get(customer.lower()) if customer else None

    # Go-Lives (2) join — authoritative go-live dates and Actual/Estimated flag.
    # Match by customer name (lowercased). SF-ID matching can be wired up later.
    gl_record = lookup_go_live(go_lives, customer)
    gl_date    = gl_record.get("go_live_date") if gl_record else None
    gl_type    = gl_record.get("go_live_type") if gl_record else None
    gl_status  = gl_record.get("status") if gl_record else None

    # Effective go-live date — Go-Lives (2) wins, Rocketlane planned is fallback.
    effective_go_live = gl_date or go_live_planned

    # Days to effective go-live
    days_to_eff_go_live = None
    if effective_go_live:
        days_to_eff_go_live = (effective_go_live - today).days

    # Exclusion flags driving milestone math
    subtype_v = get_subtype_value(p)
    excluded_pre_2026 = is_excluded_pre_2026(gl_record)
    is_phase_1 = (subtype_v == "Z2E Phase 1")
    counts_in_milestones = (
        not excluded_pre_2026
        and subtype_v in MILESTONE_PHASES
    )
    excluded_reason = None
    if excluded_pre_2026:
        excluded_reason = f"Pre-2026 actual go-live ({gl_date})"
    elif is_phase_1:
        excluded_reason = "Phase 1 (companion Phase 2 carries the ARR)"

    # Migrated proxy — use Go-Lives Actual date when available; else status.
    if gl_type and gl_type.lower() == "actual" and gl_date:
        is_live = gl_date <= today
    else:
        is_live = (status or "").lower() in LIVE_STATUSES

    return {
        "projectId":           p.get("projectId"),
        "projectName":         p.get("projectName"),
        "customer":            customer,
        "subtype":             subtype_v,
        "allSubtypeTags":      get_all_subtype_tags(p),
        "rag":                 rag,
        "status":              status,
        "goLiveDate":          effective_go_live,
        "goLiveType":          gl_type or ("Rocketlane Planned" if go_live_planned else None),
        "goLiveSheetStatus":   gl_status,
        "isLive":              is_live,
        "daysToGoLiveEff":     days_to_eff_go_live,
        "excludedFromMilestones": (not counts_in_milestones),
        "excludedReason":      excluded_reason,
        "clientSegmentation":  client_seg,
        "oppType":             opp_type,
        "forecastConfidence":  forecast_conf,
        "owner":               owner,
        "accountOwner":        account_owner,
        "oppOwner":            opp_owner,
        "startDate":           start_d,
        "dueDate":             due_d,
        "goLivePlanned":       go_live_planned,
        "daysInPhase":         days_in_phase,
        "daysToDue":           days_to_due,
        "daysToGoLive":        days_to_go_live,
        "kickoffActualDays":   kickoff_actual,
        "kickoffPlannedDays":  kickoff_planned,
        "kickoffSlipDays":     kickoff_slip,
        "ttvPlannedDays":      ttv_planned,
        "implCycleDays":       impl_cycle,
        "totalDurationDays":   total_duration,
        "modules":             n_modules,
        "modulesList":         modules_list,
        "integrations":        n_integrations,
        "integrationsList":    integrations_list,
        "trackedHours":        round(float(tracked_h), 1) if tracked_h else 0,
        "tcv":                 tcv,
        "arr":                 arr,
        "tenantDomain":        tenant_domain,
        "oppUrl":              opp_url,
        "latestStatus":        weekly_status,
    }


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
COLUMNS = [
    ("projectId",          "Project ID",            12),
    ("projectName",        "Project Name",          40),
    ("customer",           "Customer",              26),
    ("subtype",            "Phase Bucket",          18),
    ("rag",                "RAG Health",            11),
    ("status",             "Status",                14),
    ("clientSegmentation", "Client Segment",        14),
    ("oppType",            "Opp Type",              11),
    ("arr",                "ARR (Finance)",         15),
    ("tcv",                "TCV",                   13),
    ("goLiveDate",         "Go-Live Date",          14),
    ("goLiveType",         "Actual/Estimated",      18),
    ("goLiveSheetStatus",  "Sheet Status",          22),
    ("daysToGoLiveEff",    "Days to Go-Live",       14),
    ("excludedFromMilestones", "Excluded from $",   17),
    ("excludedReason",     "Exclusion Reason",      40),
    ("startDate",          "Start",                 12),
    ("dueDate",            "Due",                   12),
    ("daysInPhase",        "Days in Phase",         13),
    ("kickoffActualDays",  "Kickoff Actual (d)",    16),
    ("kickoffPlannedDays", "Kickoff Planned (d)",   17),
    ("kickoffSlipDays",    "Kickoff Slip (d)",      14),
    ("ttvPlannedDays",     "TTV Planned (d)",       15),
    ("totalDurationDays",  "Total Duration P (d)",  18),
    ("modules",            "# Modules",             10),
    ("integrations",       "# Integrations",        13),
    ("trackedHours",       "Hours Logged",          13),
    ("forecastConfidence", "Forecast Conf",         13),
    ("owner",              "PM",                    20),
    ("accountOwner",       "Account Owner",         18),
    ("oppOwner",           "Opp Owner",             18),
    ("tenantDomain",       "Tenant",                14),
    ("modulesList",        "Modules in Scope",      35),
    ("integrationsList",   "Integrations in Scope", 38),
    ("allSubtypeTags",     "All Sub-Type Tags",     22),
    ("latestStatus",       "Latest Weekly Status",  90),
    ("oppUrl",             "Opp URL",               40),
]


def style_header(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


def _col_index(key: str) -> int | None:
    """1-based column index for a given key in COLUMNS, or None if absent."""
    for i, (k, _label, _w) in enumerate(COLUMNS, start=1):
        if k == key:
            return i
    return None


def write_rows(ws, rows: list[dict]):
    rag_col = _col_index("rag")
    arr_col = _col_index("arr")
    tcv_col = _col_index("tcv")
    for r_i, row in enumerate(rows, start=2):
        for c_i, (key, _label, _w) in enumerate(COLUMNS, start=1):
            val = row.get(key)
            cell = ws.cell(row=r_i, column=c_i, value=val)
            if isinstance(val, date):
                cell.number_format = "yyyy-mm-dd"
            elif key in ("arr", "tcv") and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
        # RAG colouring
        if rag_col:
            rag = (row.get("rag") or "").lower()
            fill = None
            if rag == "red":
                fill = HEALTH_RED
            elif rag == "yellow":
                fill = HEALTH_YELLOW
            elif rag == "green":
                fill = HEALTH_GREEN
            if fill:
                ws.cell(row=r_i, column=rag_col).fill = fill


def add_data_sheet(wb, title: str, rows: list[dict]):
    ws = wb.create_sheet(title=title[:31])
    for c_i, (_, label, width) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=c_i, value=label)
        ws.column_dimensions[get_column_letter(c_i)].width = width
    style_header(ws, len(COLUMNS))
    write_rows(ws, rows)
    # Conditional format on Kickoff Slip — overdue/slipped projects in red
    slip_col_idx = _col_index("kickoffSlipDays")
    if rows and slip_col_idx:
        col_letter = get_column_letter(slip_col_idx)
        rng = f"{col_letter}2:{col_letter}{1 + len(rows)}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan",
            formula=["30"], fill=HEALTH_RED))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between",
            formula=["1", "30"], fill=HEALTH_YELLOW))
    # Conditional format on Days to Go-Live — past due in red, ≤30 days yellow
    dtg_col_idx = _col_index("daysToGoLiveEff")
    if rows and dtg_col_idx:
        col_letter = get_column_letter(dtg_col_idx)
        rng = f"{col_letter}2:{col_letter}{1 + len(rows)}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",
            formula=["0"], fill=HEALTH_RED))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between",
            formula=["0", "30"], fill=HEALTH_YELLOW))
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SNAPSHOT + TREND — writes a JSON per run, reads them all to build trend tab
# ─────────────────────────────────────────────────────────────────────────────
import json


def _is_live(row: dict) -> bool:
    """Project is considered live/migrated. Uses the per-row `isLive` flag
    set in project_row() — which prefers the Go-Lives (2) Actual date over
    Rocketlane status as the authoritative live signal."""
    return bool(row.get("isLive"))


def is_remaining_work(row: dict) -> bool:
    """True iff the project still has work left to do in the migration program.

    Excludes projects in done/wrapped/churned states (see COMPLETED_STATUSES)
    AND projects that already have an Actual go-live date in the past
    (i.e., live in production today). Pre-2026 actuals were already excluded
    upstream; this catches 2026 actuals as well.
    """
    status = (row.get("status") or "").lower()
    if status in COMPLETED_STATUSES:
        return False
    # Already gone live in production (Actual date <= today)
    if row.get("isLive"):
        return False
    return True


def dedupe_by_customer(rows: list[dict]) -> dict[str, int]:
    """Mutate rows in-place: when one customer has multiple milestone-eligible
    projects, pick a primary and mark the rest as duplicates.

    A duplicate gets `excludedFromMilestones=True` and an excludedReason of
    `"Duplicate for customer (primary: pid X)"`. ARR / counts then correctly
    sum once per customer instead of once per project.

    Primary selection rules (applied in order):
      1. Earliest goLiveDate wins (most imminent migration)
      2. Project that is already live wins over not-yet-live
      3. Highest projectId (most recently created) breaks remaining ties

    Returns {customer_name_lower: number_of_duplicates_marked} for logging.
    """
    eligible_by_customer: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        if r.get("excludedFromMilestones"):
            continue
        key = (r.get("customer") or "").lower().strip()
        if key:
            eligible_by_customer[key].append(r)

    dup_counts: dict[str, int] = {}
    for customer_key, projs in eligible_by_customer.items():
        if len(projs) <= 1:
            continue

        def sort_key(r):
            gl = r.get("goLiveDate") or date(9999, 12, 31)
            # Prefer live projects (False sorts before True with bool, so use 0/1)
            live_priority = 0 if r.get("isLive") else 1
            return (gl, live_priority, -(r.get("projectId") or 0))

        projs.sort(key=sort_key)
        primary = projs[0]
        for sec in projs[1:]:
            sec["excludedFromMilestones"] = True
            existing = sec.get("excludedReason") or ""
            note = f"Duplicate for customer (primary: pid {primary.get('projectId')})"
            sec["excludedReason"] = f"{existing}; {note}" if existing else note
        dup_counts[customer_key] = len(projs) - 1
    return dup_counts


def _counts_in_milestones(row: dict) -> bool:
    """A row counts in milestone math iff it is in MILESTONE_PHASES (i.e., not
    Phase 1) AND it isn't a pre-2026 actual go-live (excluded from the
    forward-looking analysis)."""
    return not row.get("excludedFromMilestones", False)


def compute_cadence_metrics(all_rows: list[dict]) -> dict:
    """Reduce the cohort to the headline numbers we track over time.

    Methodology rules (added 2026-04-28, snapshot schema v2):
      1. Phase 1 ARR is excluded from milestone math (companion Phase 2 carries
         the ARR; Phase 1 alone would either double-count or pre-credit work).
      2. Projects with Actual go-live before 2026-01-01 are excluded entirely
         — forward-looking tracker only.
      3. Migrated/Live status comes from the Go-Lives (2) Actual date when
         available; project status is fallback.

    Snapshots from before 2026-04-28 are v1 (status-only) and have different
    numbers — flagged in the workbook trend tab as a methodology change.
    """
    today = datetime.now().date()

    # Cohort that passes through to milestone math: Phase 2 + Not Started,
    # excluding pre-2026 actual go-lives.
    milestone_rows = [r for r in all_rows if _counts_in_milestones(r)]

    snapshot = {
        "methodology_version": 2,
        "snapshot_date": today.isoformat(),
        "cohort_size": len(all_rows),
        "cohort_size_milestone_eligible": len(milestone_rows),
        "cohort_arr_total": sum((r.get("arr") or 0) for r in milestone_rows),
        "migrated_to_date": {
            "projects": sum(1 for r in milestone_rows if _is_live(r)),
            "arr":      sum((r.get("arr") or 0) for r in milestone_rows if _is_live(r)),
        },
        "by_cutoff": {},
        "by_rag": {
            "Red":   {"projects": 0, "arr": 0.0},
            "Yellow": {"projects": 0, "arr": 0.0},
            "Green": {"projects": 0, "arr": 0.0},
            "(unset)": {"projects": 0, "arr": 0.0},
        },
        "by_phase": {},
        "exclusions": {
            "phase_1_projects": sum(1 for r in all_rows
                                    if (r.get("subtype") or "") == "Z2E Phase 1"),
            "pre_2026_actuals": sum(1 for r in all_rows
                                    if (r.get("excludedReason") or "").startswith("Pre-2026")),
            "customer_duplicates": sum(1 for r in all_rows
                                       if "Duplicate for customer" in (r.get("excludedReason") or "")),
        },
    }

    # RAG / phase rollups — only across milestone-eligible rows
    for r in milestone_rows:
        rag = (r.get("rag") or "").title() or "(unset)"
        bucket = snapshot["by_rag"].setdefault(rag, {"projects": 0, "arr": 0.0})
        bucket["projects"] += 1
        bucket["arr"] += r.get("arr") or 0
        phase = r.get("subtype") or "(none)"
        pb = snapshot["by_phase"].setdefault(phase, {"projects": 0, "arr": 0.0})
        pb["projects"] += 1
        pb["arr"] += r.get("arr") or 0

    for label, cutoff in GO_LIVE_CUTOFFS:
        eligible = [
            r for r in milestone_rows
            if r.get("goLiveDate") and r["goLiveDate"] <= cutoff
            and (r.get("status") or "").lower() != "cancelled"
        ]
        migrated_by = [r for r in eligible if _is_live(r)]
        snapshot["by_cutoff"][cutoff.isoformat()] = {
            "label":              label,
            "scheduled_projects": len(eligible),
            "scheduled_arr":      sum((r.get("arr") or 0) for r in eligible),
            "migrated_projects":  len(migrated_by),
            "migrated_arr":       sum((r.get("arr") or 0) for r in migrated_by),
        }
    return snapshot


def write_snapshot(snapshot: dict, snapshot_dir: Path = SNAPSHOT_DIR) -> Path:
    """Persist this run's snapshot. Idempotent for same-day runs (overwrites)."""
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    fname = f"z2e_cadence_{snapshot['snapshot_date']}.json"
    path = snapshot_dir / fname
    with path.open("w") as f:
        json.dump(snapshot, f, indent=2, default=str)
    return path


def load_all_snapshots(snapshot_dir: Path = SNAPSHOT_DIR) -> list[dict]:
    """Read every snapshot file in chronological order."""
    if not snapshot_dir.exists():
        return []
    snaps = []
    for p in sorted(snapshot_dir.glob("z2e_cadence_*.json")):
        try:
            snaps.append(json.loads(p.read_text()))
        except Exception as e:
            print(f"  WARN: skipping unreadable snapshot {p.name}: {e}")
    return snaps


def build_trend_sheet(wb, snapshots: list[dict]):
    """Trend tab — one row per snapshot, columns for the headline metrics.
    Easy to chart in Excel after the fact."""
    ws = wb.create_sheet(title="ARR Trend")
    ws["A1"] = "ARR Migration Trend — week-over-week"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Each row = one audit run. Compare across rows to see how "
                "scheduled and migrated ARR are moving toward the cutoff "
                "milestones. Run weekly to keep the trend tight.")
    ws["A2"].font = Font(italic=True, color="595959")

    if not snapshots:
        ws["A4"] = "No snapshots yet. Re-run the audit weekly to populate."
        return ws

    # Build columns dynamically so adding new cutoffs / RAG levels just works
    cutoff_keys = sorted({k for s in snapshots for k in s.get("by_cutoff", {}).keys()})
    rag_keys = ["Red", "Yellow", "Green", "(unset)"]

    headers = ["Snapshot Date", "Cohort", "Cohort ARR",
               "Migrated (count)", "Migrated ARR"]
    for k in cutoff_keys:
        # Pretty-print cutoff label using the first snapshot that knows it
        label = next(
            (s["by_cutoff"][k]["label"] for s in snapshots
             if k in s.get("by_cutoff", {})),
            k,
        )
        headers.extend([
            f"{label} — Sched #",
            f"{label} — Sched ARR",
            f"{label} — Migrated #",
            f"{label} — Migrated ARR",
        ])
    for rag in rag_keys:
        headers.extend([f"{rag} — #", f"{rag} — ARR"])

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(str(h)))
    ws.freeze_panes = "B5"

    for r_i, s in enumerate(snapshots, start=5):
        ws.cell(row=r_i, column=1, value=s["snapshot_date"])
        ws.cell(row=r_i, column=2, value=s.get("cohort_size", 0))
        c = ws.cell(row=r_i, column=3, value=s.get("cohort_arr_total", 0)); c.number_format = '"$"#,##0'
        ws.cell(row=r_i, column=4, value=(s.get("migrated_to_date") or {}).get("projects", 0))
        c = ws.cell(row=r_i, column=5, value=(s.get("migrated_to_date") or {}).get("arr", 0)); c.number_format = '"$"#,##0'
        col = 6
        for k in cutoff_keys:
            entry = s.get("by_cutoff", {}).get(k, {})
            ws.cell(row=r_i, column=col + 0, value=entry.get("scheduled_projects", 0))
            c = ws.cell(row=r_i, column=col + 1, value=entry.get("scheduled_arr", 0)); c.number_format = '"$"#,##0'
            ws.cell(row=r_i, column=col + 2, value=entry.get("migrated_projects", 0))
            c = ws.cell(row=r_i, column=col + 3, value=entry.get("migrated_arr", 0)); c.number_format = '"$"#,##0'
            col += 4
        for rag in rag_keys:
            entry = (s.get("by_rag") or {}).get(rag, {"projects": 0, "arr": 0.0})
            ws.cell(row=r_i, column=col, value=entry.get("projects", 0))
            c = ws.cell(row=r_i, column=col + 1, value=entry.get("arr", 0)); c.number_format = '"$"#,##0'
            col += 2

    # Show which snapshots changed deltas (last vs prior)
    if len(snapshots) >= 2:
        ws[f"A{4 + len(snapshots) + 2}"] = "Latest vs Prior Snapshot — Δ"
        ws[f"A{4 + len(snapshots) + 2}"].font = Font(bold=True)
        last, prior = snapshots[-1], snapshots[-2]
        r_i = 4 + len(snapshots) + 3
        ws.cell(row=r_i, column=1, value="Migrated ARR")
        last_v = (last.get("migrated_to_date") or {}).get("arr", 0)
        prior_v = (prior.get("migrated_to_date") or {}).get("arr", 0)
        c = ws.cell(row=r_i, column=2, value=last_v - prior_v); c.number_format = '"$"#,##0;[Red]-"$"#,##0'
        r_i += 1
        for k in cutoff_keys:
            label = last.get("by_cutoff", {}).get(k, {}).get("label", k)
            for metric_key, metric_label in [
                ("scheduled_arr", "Sched ARR"),
                ("migrated_arr", "Migrated ARR"),
            ]:
                lv = last.get("by_cutoff", {}).get(k, {}).get(metric_key, 0)
                pv = prior.get("by_cutoff", {}).get(k, {}).get(metric_key, 0)
                ws.cell(row=r_i, column=1, value=f"{label} — {metric_label}")
                c = ws.cell(row=r_i, column=2, value=lv - pv); c.number_format = '"$"#,##0;[Red]-"$"#,##0'
                r_i += 1

    return ws


def build_summary_sheet(wb, by_phase: dict[str, list[dict]],
                        unbucketed: list[dict]):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22

    today = datetime.now().strftime("%Y-%m-%d")
    ws["A1"] = f"Z2E Migration Audit — {today}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Z2E projects are NO-CHARGE retention work. Hours track team "
                "load — no project revenue. Revenue lens: ARR retention.")
    ws["A2"].font = Font(italic=True, color="595959")

    # Phase counts table — RAG-driven
    r = 4
    headers = ["Phase Bucket", "Projects", "Hours Logged", "ARR @ Risk", "Health (R / Y / G)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1
    for phase in [
        PHASE_LABEL["Z2E Phase 1"],
        PHASE_LABEL["Z2E"],
        PHASE_LABEL["Z2E - Not Started"],
    ]:
        rows = by_phase.get(phase, [])
        red = sum(1 for x in rows if (x.get("rag") or "").lower() == "red")
        yel = sum(1 for x in rows if (x.get("rag") or "").lower() == "yellow")
        grn = sum(1 for x in rows if (x.get("rag") or "").lower() == "green")
        arr_sum = sum((x.get("arr") or 0) for x in rows)
        ws.cell(row=r, column=1, value=phase)
        ws.cell(row=r, column=2, value=len(rows))
        ws.cell(row=r, column=3, value=round(sum((x.get("trackedHours") or 0) for x in rows), 1))
        c = ws.cell(row=r, column=4, value=arr_sum)
        c.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=f"{red} / {yel} / {grn}")
        r += 1

    if unbucketed:
        ws.cell(row=r, column=1, value="UNBUCKETED (review)").font = Font(bold=True, color="C00000")
        ws.cell(row=r, column=2, value=len(unbucketed))
        r += 1

    all_rows = [x for rs in by_phase.values() for x in rs]
    if not all_rows:
        return ws

    # Org-wide RAG roll-up
    r += 2
    ws.cell(row=r, column=1, value="Cohort Health Snapshot").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    r += 1
    rag_total = {"Red": 0, "Yellow": 0, "Green": 0, "(unset)": 0}
    arr_at_risk_red = 0
    arr_at_risk_yellow = 0
    for x in all_rows:
        rag = (x.get("rag") or "").title()
        if rag in rag_total:
            rag_total[rag] += 1
        else:
            rag_total["(unset)"] += 1
        if rag == "Red":
            arr_at_risk_red += x.get("arr") or 0
        elif rag == "Yellow":
            arr_at_risk_yellow += x.get("arr") or 0

    for label, count in rag_total.items():
        ws.cell(row=r, column=1, value=f"  {label}")
        ws.cell(row=r, column=2, value=count)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="ARR-at-Risk (Red projects)").font = Font(bold=True)
    c = ws.cell(row=r, column=2, value=arr_at_risk_red); c.number_format = '"$"#,##0'
    r += 1
    ws.cell(row=r, column=1, value="ARR-at-Watch (Yellow projects)").font = Font(bold=True)
    c = ws.cell(row=r, column=2, value=arr_at_risk_yellow); c.number_format = '"$"#,##0'
    r += 2

    # Top 10 ARR-at-Risk (Red)
    red_rows = [x for x in all_rows if (x.get("rag") or "").lower() == "red"]
    red_rows.sort(key=lambda x: x.get("arr") or 0, reverse=True)
    if red_rows:
        ws.cell(row=r, column=1, value="Top 10 ARR-at-Risk (Red)").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        r += 1
        for x in red_rows[:10]:
            ws.cell(row=r, column=1, value=f"  {x['customer']} — {x['projectName']}")
            c = ws.cell(row=r, column=2, value=x.get("arr") or 0); c.number_format = '"$"#,##0'
            ws.cell(row=r, column=3, value=x.get("subtype") or "")
            ws.cell(row=r, column=4, value=str(x.get("goLiveDate") or ""))
            r += 1
        r += 1

    # Strategic accounts (Client Segment = Strategic) regardless of RAG
    strategic = [x for x in all_rows if (x.get("clientSegmentation") or "") == "Strategic"]
    if strategic:
        ws.cell(row=r, column=1, value=f"Strategic Accounts ({len(strategic)})").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        r += 1
        strategic.sort(key=lambda x: x.get("arr") or 0, reverse=True)
        for x in strategic[:15]:
            ws.cell(row=r, column=1, value=f"  {x['customer']} — {x['projectName']}")
            ws.cell(row=r, column=2, value=x.get("rag") or "")
            c = ws.cell(row=r, column=3, value=x.get("arr") or 0); c.number_format = '"$"#,##0'
            ws.cell(row=r, column=4, value=x.get("subtype") or "")
            r += 1
        r += 1

    # Worst kickoff slips (actual − planned, days)
    slipped = [x for x in all_rows
               if x.get("kickoffSlipDays") is not None and x["kickoffSlipDays"] > 0]
    slipped.sort(key=lambda x: x["kickoffSlipDays"], reverse=True)
    if slipped:
        ws.cell(row=r, column=1, value="Top 10 Kickoff Slips (actual − planned, days)").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        r += 1
        for x in slipped[:10]:
            ws.cell(row=r, column=1, value=f"  {x['customer']} — {x['projectName']}")
            ws.cell(row=r, column=2, value=int(x["kickoffSlipDays"]))
            ws.cell(row=r, column=3, value=x.get("rag") or "")
            ws.cell(row=r, column=4, value=x.get("subtype") or "")
            r += 1

    return ws


def build_go_live_cadence_sheet(wb, all_rows: list[dict]):
    """ARR / project counts by Go-Live target cutoff. Tracking metric tab.

    For each cutoff date in GO_LIVE_CUTOFFS, shows:
      - Projects with Go-Live target on or before the cutoff (cumulative)
      - Sum of ARR for those projects (from CSV)
      - Split: already past Go-Live (target date < today) vs upcoming
      - Then a project-level list under each cutoff
    """
    ws = wb.create_sheet(title="Go-Live Cadence")
    today = datetime.now().date()

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    ws["A1"] = "Go-Live Cadence — ARR Tracking"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Aggregate ARR scheduled to go live by each cutoff. "
                "Tracking metric for the Z2E retention program.")
    ws["A2"].font = Font(italic=True, color="595959")
    ws["A3"] = ("Sources: Go-Live Target = `eDisc: Go Live - Planned` (Rocketlane). "
                "ARR = Customers tab CSV, Column C 'ARR from Finance (Jan'26)'. "
                "Salesforce ARR is unreliable and intentionally excluded.")
    ws["A3"].font = Font(italic=True, color="595959", size=9)

    # ─── Rollup table — Scheduled vs Migrated by each cutoff ─────────────
    # Methodology (2026-04-28):
    #   - Phase 1 projects EXCLUDED (companion Phase 2 carries the ARR)
    #   - Pre-2026 actual go-lives EXCLUDED (forward-looking only)
    #   - Migrated = Go-Lives (2) Actual date <= today (status as fallback)
    milestone_rows = [r for r in all_rows if _counts_in_milestones(r)]
    excluded_phase1 = sum(1 for r in all_rows
                          if (r.get("subtype") or "") == "Z2E Phase 1")
    excluded_pre26 = sum(1 for r in all_rows
                         if (r.get("excludedReason") or "").startswith("Pre-2026"))
    excluded_dup = sum(1 for r in all_rows
                       if "Duplicate for customer" in (r.get("excludedReason") or ""))

    r = 5
    ws.cell(row=r, column=1,
            value=(f"Eligible cohort: {len(milestone_rows)} projects  "
                   f"(excluded: {excluded_phase1} Phase 1, "
                   f"{excluded_pre26} pre-2026 actuals, "
                   f"{excluded_dup} customer duplicates)")
            ).font = Font(italic=True, color="595959")
    r += 2

    headers = ["Cutoff", "Scheduled #", "Scheduled ARR",
               "Migrated # (live)", "Migrated ARR",
               "Remaining ARR", "% Migrated", "% with ARR data"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1

    cutoff_buckets: dict[str, list[dict]] = {}
    for label, cutoff in GO_LIVE_CUTOFFS:
        eligible = [x for x in milestone_rows
                    if x.get("goLiveDate") and x["goLiveDate"] <= cutoff
                    and (x.get("status") or "").lower() != "cancelled"]
        migrated = [x for x in eligible if _is_live(x)]
        with_arr = [x for x in eligible if x.get("arr")]
        sched_arr = sum((x.get("arr") or 0) for x in eligible)
        migrated_arr = sum((x.get("arr") or 0) for x in migrated)
        remaining_arr = sched_arr - migrated_arr
        pct_migrated = (100 * migrated_arr / sched_arr) if sched_arr else 0
        pct_arr = (100 * len(with_arr) / len(eligible)) if eligible else 0

        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=len(eligible))
        c = ws.cell(row=r, column=3, value=sched_arr); c.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=len(migrated))
        c = ws.cell(row=r, column=5, value=migrated_arr); c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=6, value=remaining_arr); c.number_format = '"$"#,##0'
        ws.cell(row=r, column=7, value=f"{pct_migrated:.1f}%")
        ws.cell(row=r, column=8, value=f"{pct_arr:.0f}%")
        r += 1

        cutoff_buckets[label] = eligible

    # ─── Projects per cutoff ───────────────────────────────────────────────
    # We list each project under the *earliest* cutoff bucket it falls into,
    # so each project appears once. Then a "Beyond 12/31" section catches
    # everything else with a Go-Live target.
    listed_pids: set = set()

    for label, cutoff in GO_LIVE_CUTOFFS:
        rows = sorted(
            [x for x in cutoff_buckets[label] if x.get("projectId") not in listed_pids],
            key=lambda x: (x.get("goLiveDate") or date(9999, 1, 1), -(x.get("arr") or 0)),
        )
        r += 2
        ws.cell(row=r, column=1, value=f"{label} — {len(rows)} projects").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        r += 1
        sub_headers = ["Customer / Project", "Go-Live Target", "RAG", "ARR", "Phase", "Status"]
        for c, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True)
        r += 1
        for x in rows:
            ws.cell(row=r, column=1, value=f"{x['customer']} — {x['projectName']}")
            c = ws.cell(row=r, column=2, value=x.get("goLiveDate"))
            c.number_format = "yyyy-mm-dd"
            rag_cell = ws.cell(row=r, column=3, value=x.get("rag") or "")
            rag = (x.get("rag") or "").lower()
            if rag == "red":
                rag_cell.fill = HEALTH_RED
            elif rag == "yellow":
                rag_cell.fill = HEALTH_YELLOW
            elif rag == "green":
                rag_cell.fill = HEALTH_GREEN
            c = ws.cell(row=r, column=4, value=x.get("arr") or 0)
            c.number_format = '"$"#,##0'
            ws.cell(row=r, column=5, value=x.get("subtype") or "")
            ws.cell(row=r, column=6, value=x.get("status") or "")
            listed_pids.add(x.get("projectId"))
            r += 1

    # Beyond final cutoff — projects with Go-Live target > last cutoff
    last_cutoff = GO_LIVE_CUTOFFS[-1][1]
    beyond = [x for x in all_rows
              if x.get("goLiveDate") and x["goLiveDate"] > last_cutoff
              and (x.get("status") or "").lower() not in ("cancelled", "completed", "closeout")]
    if beyond:
        beyond.sort(key=lambda x: x["goLiveDate"])
        r += 2
        arr_beyond = sum((x.get("arr") or 0) for x in beyond)
        ws.cell(row=r, column=1, value=f"Beyond {last_cutoff.strftime('%-m/%-d/%Y')} — {len(beyond)} projects").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        c = ws.cell(row=r, column=4, value=arr_beyond); c.number_format = '"$"#,##0'
        r += 1
        for c, h in enumerate(["Customer / Project", "Go-Live Target", "RAG", "ARR", "Phase", "Status"], 1):
            ws.cell(row=r, column=c, value=h).font = Font(bold=True)
        r += 1
        for x in beyond:
            ws.cell(row=r, column=1, value=f"{x['customer']} — {x['projectName']}")
            c = ws.cell(row=r, column=2, value=x.get("goLiveDate")); c.number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=3, value=x.get("rag") or "")
            c = ws.cell(row=r, column=4, value=x.get("arr") or 0); c.number_format = '"$"#,##0'
            ws.cell(row=r, column=5, value=x.get("subtype") or "")
            ws.cell(row=r, column=6, value=x.get("status") or "")
            r += 1

    # Projects with NO Go-Live target — flag for PMs
    no_target = [x for x in all_rows
                 if not x.get("goLiveDate")
                 and (x.get("status") or "").lower() not in ("cancelled", "completed", "closeout")]
    if no_target:
        r += 2
        ws.cell(row=r, column=1, value=f"NO Go-Live Target Set — {len(no_target)} projects (PMs to populate)").font = Font(bold=True, color="C00000")
        r += 1
        for c, h in enumerate(["Customer / Project", "Phase", "Status", "Days in Phase"], 1):
            ws.cell(row=r, column=c, value=h).font = Font(bold=True)
        r += 1
        no_target.sort(key=lambda x: x.get("daysInPhase") or 0, reverse=True)
        for x in no_target:
            ws.cell(row=r, column=1, value=f"{x['customer']} — {x['projectName']}")
            ws.cell(row=r, column=2, value=x.get("subtype") or "")
            ws.cell(row=r, column=3, value=x.get("status") or "")
            ws.cell(row=r, column=4, value=x.get("daysInPhase"))
            r += 1

    return ws


def build_resource_sheet(wb, te_by_proj: dict[int, list[dict]],
                          rows_by_phase: dict[str, list[dict]]):
    """Hours by user across the Z2E book, with phase split."""
    ws = wb.create_sheet(title="Resource Load")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    pid_to_phase: dict[int, str] = {}
    for phase, rs in rows_by_phase.items():
        for r in rs:
            pid_to_phase[r["projectId"]] = phase

    user_phase_minutes: dict[tuple[str, str], int] = defaultdict(int)
    user_total_minutes: dict[str, int] = defaultdict(int)
    for pid, entries in te_by_proj.items():
        phase = pid_to_phase.get(pid, "Other")
        for e in entries:
            u = e.get("user") or {}
            name = (u.get("firstName", "") + " " + u.get("lastName", "")).strip() or \
                   u.get("emailId", "") or "Unknown"
            mins = e.get("minutes") or 0
            user_phase_minutes[(name, phase)] += mins
            user_total_minutes[name] += mins

    headers = ["User", "Phase 1 hrs", "Phase 2 hrs", "Not-Started hrs", "Total hrs"]
    for c_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=c_i, value=h)
    style_header(ws, len(headers))

    sorted_users = sorted(user_total_minutes.items(),
                          key=lambda kv: kv[1], reverse=True)
    for r_i, (user, _) in enumerate(sorted_users, start=2):
        ws.cell(row=r_i, column=1, value=user)
        ws.cell(row=r_i, column=2,
                value=round(user_phase_minutes[(user, PHASE_LABEL["Z2E Phase 1"])] / 60, 1))
        ws.cell(row=r_i, column=3,
                value=round(user_phase_minutes[(user, PHASE_LABEL["Z2E"])] / 60, 1))
        ws.cell(row=r_i, column=4,
                value=round(user_phase_minutes[(user, PHASE_LABEL["Z2E - Not Started"])] / 60, 1))
        ws.cell(row=r_i, column=5, value=round(user_total_minutes[user] / 60, 1))


def build_milestones_at_risk_sheet(wb, projects: list[dict],
                                    ms_by_pid: dict[int, list[dict]]):
    ws = wb.create_sheet(title="Milestones at Risk")
    headers = ["Project", "Customer", "Milestone", "Due", "Days from Today", "Status"]
    for c_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=c_i, value=h)
        ws.column_dimensions[get_column_letter(c_i)].width = 22 if c_i in (1,3) else 16
    style_header(ws, len(headers))

    today = datetime.now().date()
    horizon = today + timedelta(days=14)
    rows = []
    pid_to_meta = {p.get("projectId"): p for p in projects}
    for pid, ms in ms_by_pid.items():
        p = pid_to_meta.get(pid, {})
        for m in ms:
            d = parse_iso_date(m.get("dueDate") or m.get("plannedDate"))
            status = (m.get("status") or {}).get("label", "")
            if not d or "done" in status.lower() or "complete" in status.lower():
                continue
            if d <= horizon:
                rows.append((p.get("projectName") or "",
                             (p.get("customer") or {}).get("companyName") or "",
                             m.get("name") or m.get("phaseName") or "",
                             d, (d - today).days, status))
    rows.sort(key=lambda r: r[3])
    for r_i, row in enumerate(rows, start=2):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            if isinstance(val, date):
                cell.number_format = "yyyy-mm-dd"
        if row[4] < 0:  # overdue
            ws.cell(row=r_i, column=5).fill = HEALTH_RED
        elif row[4] <= 7:
            ws.cell(row=r_i, column=5).fill = HEALTH_YELLOW


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Z2E migration program audit.")
    ap.add_argument("--since", default=DEFAULT_SINCE,
                    help=f"Time-entry window start (default: {DEFAULT_SINCE}, "
                         f"only used with --include-time-entries)")
    ap.add_argument("--include-time-entries", action="store_true",
                    help="Also pull granular time entries for the Resource Load "
                         "tab and 30-day burn trend. OFF by default — project "
                         "list endpoint already returns trackedHours / budget "
                         "rollups per project, which is enough for project-level "
                         "burn analysis. Use this flag only when you need a "
                         "by-person hours breakdown.")
    ap.add_argument("--arr-csv", "--arr-file", dest="arr_csv", default=None,
                    help="ARR source file (CSV or XLSX) — Customers tab from "
                         "the Z2E tracking spreadsheet. Looks for an 'ARR from "
                         "Finance' column. Optional, but Go-Live Cadence ARR "
                         "metrics will be empty without it.")
    ap.add_argument("--no-snapshot", action="store_true",
                    help="Skip writing today's snapshot to z2e_snapshots/. "
                         "Default is to snapshot every run for trend tracking.")
    ap.add_argument("--remaining-only", action="store_true",
                    help="Filter cohort to projects with work remaining "
                         "(excludes Completed, Closeout, Cancelled, Hypercare, "
                         "Partially Live, and any project already past Actual "
                         "go-live). Writes a separate workbook "
                         "Z2E_Remaining_Audit_<date>.xlsx and snapshots into "
                         "z2e_remaining_snapshots/ so the trend stays "
                         "consistent.")
    ap.add_argument("--out", default=None,
                    help="Output xlsx path (default: Z2E_Migration_Audit_<today>.xlsx)")
    args = ap.parse_args()

    today_str = datetime.now().strftime("%Y-%m-%d")
    if args.remaining_only:
        default_name = f"Z2E_Remaining_Audit_{today_str}.xlsx"
        active_snap_dir = REMAINING_SNAPSHOT_DIR
    else:
        default_name = f"Z2E_Migration_Audit_{today_str}.xlsx"
        active_snap_dir = SNAPSHOT_DIR
    out_path = args.out or str(DEFAULT_OUTPUT_DIR / default_name)

    print(f"[{today_str}] Z2E migration audit starting")

    # 1. Locate the Z2E sub-type fieldId
    print(f"[1/5] Resolving fieldId for '{SUBTYPE_FIELD_LABEL}'...")
    field_id = find_field_id(SUBTYPE_FIELD_LABEL)
    if not field_id:
        sys.exit(f"FATAL: '{SUBTYPE_FIELD_LABEL}' field not found via /fields")
    print(f"      fieldId = {field_id}")

    # 2. Pull cohort
    print("[2/5] Pulling Z2E project cohort...")
    projects, calls = fetch_z2e_projects(field_id)
    print(f"      {len(projects)} projects in {calls} API call(s)")

    # 3. Bucket by sub-type, with name-fallback for any field gaps
    by_phase_pids: dict[str, list[int]] = defaultdict(list)
    by_phase_proj: dict[str, list[dict]] = defaultdict(list)
    unbucketed: list[dict] = []
    field_hit = 0
    field_miss_name_hit = 0
    raw_subtype_samples: list[str] = []
    for p in projects:
        sub = get_subtype_value(p)
        if sub:
            field_hit += 1
            phase = PHASE_LABEL[sub]
        else:
            field_miss_name_hit += 1
            # Capture a few samples of what the field actually looks like
            if len(raw_subtype_samples) < 3:
                for f in p.get("fields") or []:
                    if f.get("fieldLabel") == SUBTYPE_FIELD_LABEL:
                        raw_subtype_samples.append(
                            f"  pid={p.get('projectId')} field={f}"
                        )
                        break
            name = (p.get("projectName") or "").lower()
            if "phase 1" in name:
                phase = PHASE_LABEL["Z2E Phase 1"]
            elif "not started" in name or "not-started" in name:
                phase = PHASE_LABEL["Z2E - Not Started"]
            elif "z2e" in name or "zapproved" in name:
                phase = PHASE_LABEL["Z2E"]
            else:
                unbucketed.append(p)
                continue
        by_phase_proj[phase].append(p)
        by_phase_pids[phase].append(p.get("projectId"))
    print(f"      bucketing: {field_hit} via sub-type field, "
          f"{field_miss_name_hit} via name fallback")
    if raw_subtype_samples:
        print("      sample sub-type field shape from a name-fallback project:")
        for s in raw_subtype_samples:
            print(s[:240])
    for phase, items in by_phase_proj.items():
        print(f"      {phase}: {len(items)}")
    if unbucketed:
        print(f"      UNBUCKETED: {len(unbucketed)} (review names)")

    # 4. Milestones — currently disabled (endpoint not yet verified).
    print("[3/5] Skipping milestones (endpoint pending verification — see source).")
    ms_by_pid: dict[int, list[dict]] = {p.get("projectId"): [] for p in projects}

    # 5. Time entries — OFF by default.
    #    Project-level trackedHours / budgetedHours / percentageBudgetedHoursConsumed
    #    rollups are already on the project payload from step [2/5]. Granular
    #    time entries are only needed for the by-person Resource Load tab and
    #    the 30-day burn trend. Skip unless explicitly requested.
    te_by_pid: dict[int, list[dict]] = {}
    if args.include_time_entries:
        print(f"[4/5] Pulling time entries since {args.since} (per-project, 4 workers)...")
        pids = [p.get("projectId") for p in projects if p.get("projectId")]
        try:
            te_by_pid, te_calls = fetch_time_entries_per_project(
                pids, since_date=args.since, max_workers=4
            )
            total_entries = sum(len(v) for v in te_by_pid.values())
            print(f"      {total_entries} entries on {len(te_by_pid)} projects "
                  f"in {te_calls} call(s)")
        except Exception as e:
            print(f"      WARN time-entries: {str(e)[:200]}")
            te_by_pid = {}
    else:
        print("[4/5] Skipping time-entry pull. "
              "Using project-level trackedHours rollup. "
              "Pass --include-time-entries for by-person Resource Load + 30d burn trend.")

    # 6. ARR join + Go-Lives (2) join — both come from the same xlsx
    arr_lookup: dict[str, float] = {}
    go_lives: dict[str, dict] = {}
    if args.arr_csv:
        print(f"      Loading ARR from {args.arr_csv}...")
        arr_lookup = load_arr_file(args.arr_csv)
        print(f"      {len(arr_lookup)} customers with ARR")
        # Same workbook → load Go-Lives (2) tab for actual/estimated dates
        go_lives = load_go_lives_xlsx(args.arr_csv)
        if go_lives:
            n_actual = sum(1 for v in go_lives.values()
                           if (v.get("go_live_type") or "").lower() == "actual")
            print(f"      {len([k for k in go_lives if not k.startswith('sfid:')])} "
                  f"customers in Go-Lives (2) ({n_actual} Actual)")

    # 7. Build per-project derived rows, grouped by phase
    rows_by_phase: dict[str, list[dict]] = defaultdict(list)
    for phase, ps in by_phase_proj.items():
        for p in ps:
            row = project_row(p,
                              ms_by_pid.get(p.get("projectId"), []),
                              te_by_pid.get(p.get("projectId"), []),
                              arr_lookup,
                              go_lives)
            rows_by_phase[phase].append(row)
    unbucketed_rows = [project_row(p,
                                    ms_by_pid.get(p.get("projectId"), []),
                                    te_by_pid.get(p.get("projectId"), []),
                                    arr_lookup,
                                    go_lives)
                       for p in unbucketed]

    # Customer-level dedup — when one customer has multiple Rocketlane projects
    # in the milestone cohort, only the primary carries the ARR. Secondaries
    # are flagged as duplicates so they don't double-count.
    all_built = [r for rs in rows_by_phase.values() for r in rs] + unbucketed_rows

    # Remaining-only filter — applied BEFORE dedup so the primary picked is
    # one of the still-active projects, not a completed sibling.
    if args.remaining_only:
        before = len(all_built)
        all_built = [r for r in all_built if is_remaining_work(r)]
        dropped = before - len(all_built)
        # Rebuild rows_by_phase to reflect the filtered set so workbook tabs
        # only show remaining projects.
        kept_pids = {r["projectId"] for r in all_built}
        rows_by_phase = {
            phase: [r for r in rs if r["projectId"] in kept_pids]
            for phase, rs in rows_by_phase.items()
        }
        unbucketed_rows = [r for r in unbucketed_rows
                           if r["projectId"] in kept_pids]
        print(f"      --remaining-only: dropped {dropped} done/churned/live "
              f"projects; {len(all_built)} remain in cohort")

    dup_counts = dedupe_by_customer(all_built)
    if dup_counts:
        n_dups = sum(dup_counts.values())
        print(f"      deduped {n_dups} secondary project(s) across "
              f"{len(dup_counts)} customer(s):")
        for cust, n in sorted(dup_counts.items(), key=lambda kv: -kv[1])[:10]:
            print(f"        - {cust}: {n} duplicate(s) marked")

    # Print exclusion summary so the run is self-documenting
    n_excl_p1 = sum(1 for r in all_built if (r.get("subtype") or "") == "Z2E Phase 1")
    n_excl_pre = sum(1 for r in all_built
                     if (r.get("excludedReason") or "").startswith("Pre-2026"))
    n_excl_dup = sum(1 for r in all_built
                     if "Duplicate for customer" in (r.get("excludedReason") or ""))
    n_eligible = sum(1 for r in all_built if not r.get("excludedFromMilestones"))
    print(f"      milestone-eligible cohort: {n_eligible} projects "
          f"(excluded: {n_excl_p1} Phase 1, {n_excl_pre} pre-2026 actuals, "
          f"{n_excl_dup} customer duplicates)")
    n_with_gl = sum(1 for r in all_built if r.get("goLiveDate"))
    n_actual_live = sum(1 for r in all_built
                        if (r.get("goLiveType") or "").lower() == "actual")
    print(f"      {n_with_gl} projects have a go-live date "
          f"({n_actual_live} Actual from Go-Lives (2) tab)")

    # 8. Workbook
    print(f"[5/5] Writing {out_path}...")
    wb = Workbook()
    wb.remove(wb.active)
    build_summary_sheet(wb, rows_by_phase, unbucketed_rows)
    add_data_sheet(wb, "Phase 1 — Tech Integration",
                   rows_by_phase.get(PHASE_LABEL["Z2E Phase 1"], []))
    add_data_sheet(wb, "Phase 2 — Migration & Ramp",
                   rows_by_phase.get(PHASE_LABEL["Z2E"], []))
    add_data_sheet(wb, "Not Started Queue",
                   rows_by_phase.get(PHASE_LABEL["Z2E - Not Started"], []))
    if unbucketed_rows:
        add_data_sheet(wb, "Unbucketed (review)", unbucketed_rows)
    # Go-Live Cadence tracking tab — ARR by 6/30, 9/30, 12/31 cutoff
    all_rows = [r for rs in rows_by_phase.values() for r in rs] + unbucketed_rows
    build_go_live_cadence_sheet(wb, all_rows)

    # Snapshot — append today's metrics to the historical trend.
    # Remaining-only mode writes to its own snapshot dir so the trend doesn't
    # mix two different cohort definitions.
    if not args.no_snapshot:
        try:
            snapshot = compute_cadence_metrics(all_rows)
            snapshot["mode"] = "remaining_only" if args.remaining_only else "full"
            snap_path = write_snapshot(snapshot, snapshot_dir=active_snap_dir)
            print(f"      snapshot written: {snap_path}")
        except Exception as e:
            print(f"      WARN: snapshot write failed: {e}")

    # Trend tab — reads every snapshot from this mode's directory
    snapshots = load_all_snapshots(snapshot_dir=active_snap_dir)
    if snapshots:
        build_trend_sheet(wb, snapshots)
        print(f"      trend tab built from {len(snapshots)} snapshot(s)")

    # Milestones tab skipped — endpoint pending verification
    if te_by_pid:
        build_resource_sheet(wb, te_by_pid, rows_by_phase)
    # Raw projects (one row per project, all derived metrics)
    add_data_sheet(wb, "Raw — All Projects",
                   [r for rs in rows_by_phase.values() for r in rs] + unbucketed_rows)

    wb.save(out_path)
    print(f"Done.  {out_path}")


if __name__ == "__main__":
    main()
